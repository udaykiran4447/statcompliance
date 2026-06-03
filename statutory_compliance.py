# ─────────────────────────────────────────────────────────────────────────────
#  STATUTORY COMPLIANCE SUITE  –  Combined App  (10 tools)
# ─────────────────────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import re, io, os, json, math, zipfile, warnings
from io import BytesIO
from pathlib import Path
from datetime import date, datetime
import openpyxl
warnings.filterwarnings("ignore")

import pdfplumber
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta

try:
    import xlrd
except ImportError:
    xlrd = None
try:
    import numpy as np
except ImportError:
    import types; np = types.SimpleNamespace(isnan=lambda x: x!=x)

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Statutory Compliance Suite",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── GLOBAL CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700&family=Source+Sans+3:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Source Sans 3',sans-serif;}
.stApp{background:#f4f1eb;}
.block-container{padding:1.6rem 2.2rem 3rem;max-width:1440px;}

/* Sidebar */
section[data-testid="stSidebar"]{
  background:linear-gradient(180deg,#0b1f3a 0%,#122849 60%,#0e2040 100%);
  border-right:3px solid #c9a84c;min-width:270px;}
section[data-testid="stSidebar"] *{color:#e8dfc8 !important;}
section[data-testid="stSidebar"] hr{border-color:rgba(201,168,76,.35) !important;}
.sb-logo{text-align:center;padding:22px 12px 14px;
  border-bottom:1px solid rgba(201,168,76,.3);margin-bottom:6px;}
.sb-logo .sb-title{font-family:'Playfair Display',serif;font-size:1.15rem;
  font-weight:700;color:#c9a84c !important;letter-spacing:.03em;line-height:1.3;}
.sb-logo .sb-sub{font-size:.72rem;color:#8fa8c8 !important;
  letter-spacing:.12em;text-transform:uppercase;margin-top:3px;}
div[data-testid="stSidebar"] .stButton>button{
  width:100%;text-align:left !important;background:transparent !important;
  border:none !important;border-left:3px solid transparent !important;
  border-radius:0 6px 6px 0 !important;color:#c8d8e8 !important;
  font-size:.85rem !important;font-weight:500 !important;
  padding:9px 14px !important;margin:1px 0 !important;transition:all .2s !important;}
div[data-testid="stSidebar"] .stButton>button:hover{
  background:rgba(201,168,76,.12) !important;
  border-left-color:#c9a84c !important;color:#f5e6b0 !important;}

/* Page header */
.page-header{background:linear-gradient(135deg,#0b1f3a 0%,#1a3a5c 60%,#0e2040 100%);
  border-radius:14px;padding:26px 32px;margin-bottom:24px;
  border-bottom:4px solid #c9a84c;box-shadow:0 6px 28px rgba(11,31,58,.22);}
.page-header h1{font-family:'Playfair Display',serif;font-size:1.85rem;
  font-weight:700;color:#f5e6b0;margin:0 0 4px 0;}
.page-header p{font-size:.92rem;color:#8fb8d8;margin:0;}

/* Home cards */
.home-card{background:white;border-radius:12px;padding:22px 20px 18px;
  border:1px solid #d6c9a8;border-top:4px solid #c9a84c;
  box-shadow:0 2px 12px rgba(11,31,58,.08);transition:all .25s;}
.home-card:hover{transform:translateY(-3px);
  box-shadow:0 8px 24px rgba(11,31,58,.15);border-top-color:#0b1f3a;}
.home-card .hc-icon{font-size:2.1rem;margin-bottom:10px;}
.home-card .hc-title{font-family:'Playfair Display',serif;font-size:1rem;
  font-weight:700;color:#0b1f3a;margin-bottom:5px;}
.home-card .hc-desc{font-size:.78rem;color:#6b7280;line-height:1.45;}

/* Step headers / info */
.step-header{font-size:1rem;font-weight:600;color:#0b1f3a;background:#e9e2d0;
  padding:.5rem 1rem;border-left:4px solid #c9a84c;border-radius:4px;
  margin-bottom:.8rem;margin-top:.4rem;}
.info-box{background:#f0f4fa;border:1px solid #c2d9f7;border-radius:6px;
  padding:.65rem 1rem;font-size:.87rem;color:#1a3c5e;}
.success-box{background:#e6f9f0;border:1px solid #a3dfc4;border-radius:6px;
  padding:.7rem 1rem;font-size:.93rem;color:#1a5c3a;}

/* Metric / stat cards */
.metric-card{background:white;border-radius:10px;padding:16px 18px;
  text-align:center;border:1px solid #d6c9a8;border-top:3px solid #c9a84c;}
.metric-value{font-size:1.9rem;font-weight:700;color:#0b1f3a;}
.metric-label{font-size:.78rem;color:#6b7280;margin-top:3px;}
.stat-card{background:white;border:1px solid #d6c9a8;border-radius:10px;
  padding:1.1rem 1.4rem;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.06);}
.stat-card .value{font-size:1.85rem;font-weight:700;color:#0b1f3a;}
.stat-card .label{font-size:.82rem;color:#666;margin-top:.2rem;}

/* Primary button */
.stButton>button[kind="primary"],button[data-testid="baseButton-primary"]{
  background:linear-gradient(135deg,#0b1f3a,#1a3a5c) !important;
  color:#f5e6b0 !important;border:none !important;border-radius:8px !important;
  font-weight:600 !important;}
.stDownloadButton>button{
  background:linear-gradient(135deg,#14532d,#166534) !important;
  color:white !important;border:none !important;border-radius:8px !important;font-weight:600 !important;}

/* File uploader */
div[data-testid="stFileUploader"]{
  border:2px dashed #c9a84c !important;border-radius:12px !important;
  background:#faf7f0 !important;padding:6px !important;}

/* Tabs */
.stTabs [data-baseweb="tab-list"]{gap:2px;background:#e9e2d0;
  border-radius:8px 8px 0 0;padding:4px 4px 0;}
.stTabs [data-baseweb="tab"]{font-size:.85rem;font-weight:500;
  border-radius:6px 6px 0 0;padding:7px 18px;color:#374151;}
.stTabs [aria-selected="true"]{background:white !important;
  color:#0b1f3a !important;border-bottom:2px solid #c9a84c !important;}

/* IndAS callouts */
.callout{background:#EFF6FF;border-left:3px solid #3B82F6;border-radius:0 6px 6px 0;
  padding:9px 13px;font-size:12px;color:#1E40AF;margin:4px 0 10px;}
.camber{background:#FFFBEB;border-left:3px solid #F59E0B;border-radius:0 6px 6px 0;
  padding:9px 13px;font-size:12px;color:#92400E;margin:4px 0 10px;}

/* Sheet pill */
.sheet-pill{display:inline-block;background:#e9e2d0;color:#0b1f3a;
  border-radius:20px;padding:.2rem .8rem;font-size:.78rem;margin:.2rem;font-weight:500;}

label{font-size:12px !important;font-weight:600 !important;color:#374151 !important;}
</style>
""", unsafe_allow_html=True)

# ── SESSION STATE ─────────────────────────────────────────────────────────────
PAGES = [
    ("🏠",  "Home"),
    ("🧾",  "GSTR-1 JSON Converter"),
    ("📊",  "GSTR-2A Combiner"),
    ("📋",  "GSTR-2B Consolidator"),
    ("📦",  "E-Way Bill Consolidator"),
    ("📂",  "Excel Consolidator"),
    ("🔄",  "E-Invoice Consolidator"),
    ("💸",  "TDS Challan Extractor"),
    ("🏭",  "EPF Challan Consolidator"),
    ("🏥",  "ESI Challan Extractor"),
    ("📐",  "Ind AS 116 – Lease"),
]
if "page" not in st.session_state:
    st.session_state.page = "Home"

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div class='sb-logo'>
      <div style='font-size:2.4rem;'>⚖️</div>
      <div class='sb-title'>Statutory<br>Compliance Suite</div>
      <div class='sb-sub'>10 Tools · One Platform</div>
    </div>""", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    for icon, name in PAGES:
        if st.button(f"{icon}  {name}", key=f"nav_{name}", use_container_width=True):
            st.session_state.page = name
            st.rerun()
    st.markdown("<br><hr>", unsafe_allow_html=True)
    st.markdown(
        "<div style='font-size:.72rem;color:#4a6a8a;text-align:center;padding:0 8px;'>"
        "All processing is local. No data sent to any third party.</div>",
        unsafe_allow_html=True)

ACTIVE = st.session_state.page

# ── HELPERS ───────────────────────────────────────────────────────────────────
def page_header(icon, title, subtitle=""):
    st.markdown(f"<div class='page-header'><h1>{icon} {title}</h1><p>{subtitle}</p></div>",
                unsafe_allow_html=True)
def step_header(text):
    st.markdown(f"<div class='step-header'>{text}</div>", unsafe_allow_html=True)
def info_box(text):
    st.markdown(f"<div class='info-box'>{text}</div>", unsafe_allow_html=True)

# ── HOME ──────────────────────────────────────────────────────────────────────
if ACTIVE == "Home":
    page_header("⚖️", "Statutory Compliance Suite",
                "A unified platform for statutory compliance tools — GST · TDS · PF · ESI · E-Way Bills · Lease Accounting")
    tools = [
        ("🧾","GSTR-1 JSON Converter",     "Convert GSTR-1 JSON / ZIP files into a multi-sheet Excel workbook."),
        ("📊","GSTR-2A Combiner",           "Combine multiple GSTR-2A Excel files sheet-wise, removing subtotals."),
        ("📋","GSTR-2B Consolidator",       "Consolidate multiple GSTR-2B files across periods into one report."),
        ("📦","E-Way Bill Consolidator",    "Merge Chattisgarh E-Way Bill HTML-XLS portal exports."),
        ("📂","Excel Consolidator",         "Merge XLS, XLSX, XLSM & CSV files into one output."),
        ("🔄","E-Invoice Consolidator",     "Upload multiple Excel files, pick sheets & header row, download consolidated."),
        ("💸","TDS Challan Extractor",      "Extract ITNS 281 TDS challan data from PDFs into Excel."),
        ("🏭","EPF Challan Consolidator",   "Parse EPF Combined Challan PDFs into a formatted Excel report."),
        ("🏥","ESI Challan Extractor",      "Extract ESIC challan data from PDFs and export to Excel."),
        ("📐","Ind AS 116 – Lease",         "Compute lease liability amortisation schedules per Ind AS 116."),
    ]
    cols = st.columns(3)
    for i, (icon, name, desc) in enumerate(tools):
        with cols[i % 3]:
            st.markdown(f"""<div class='home-card'>
              <div class='hc-icon'>{icon}</div>
              <div class='hc-title'>{name}</div>
              <div class='hc-desc'>{desc}</div></div>""", unsafe_allow_html=True)
            if st.button("Open →", key=f"hb_{name}"):
                st.session_state.page = name; st.rerun()
    st.markdown("<br>", unsafe_allow_html=True)
    st.info("👈 Use the sidebar to navigate. Collapse it with the **›** arrow on its edge.")

# ── GSTR-1 JSON Converter ───────────────────────────────────────
elif ACTIVE == "GSTR-1 JSON Converter":
    page_header("🧾", "GSTR-1 JSON → Excel Converter", "Upload one or more GSTR-1 JSON / ZIP files and consolidate into a single Excel workbook.")
    """
    GSTR-1 JSON → Excel Converter  (Multi-file / Consolidated)
    Run: streamlit run gstr1_converter.py
    """


    # ─── Palette & styles ─────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill('solid', start_color='1F4E79')
    HDR2_FILL  = PatternFill('solid', start_color='2E75B6')
    ALT_FILL   = PatternFill('solid', start_color='EBF3FB')
    WHITE_FILL = PatternFill('solid', start_color='FFFFFF')
    TITLE_FILL = PatternFill('solid', start_color='DEEAF1')
    SRC_FILL   = PatternFill('solid', start_color='FFF2CC')   # yellow tint for source col header
    HDR_FONT   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    SRC_HDR_FONT = Font(bold=True, color='7F6000', name='Arial', size=10)
    CELL_FONT  = Font(name='Arial', size=9)
    TITLE_FONT = Font(bold=True, name='Arial', size=11, color='1F4E79')
    _thin      = Side(style='thin', color='B8CCE4')
    BORDER     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    FILE_COL_W = 30   # width for the "Source File" column


    def _style_hdr(ws, row, n_cols):
        """Style header row; col-1 is the Source File column (yellow), rest dark blue."""
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = SRC_FILL if c == 1 else HDR_FILL
            cell.font = SRC_HDR_FONT if c == 1 else HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER


    def _style_sum_hdr(ws, row, n_cols):
        """Style summary sub-header (all same blue)."""
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR2_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER


    def _style_data(ws, row, n_cols, alt=False):
        fill = ALT_FILL if alt else WHITE_FILL
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal='right' if isinstance(cell.value, (int, float)) else 'left'
            )


    def _fmt_num(ws, row, col_indices):
        for c in col_indices:
            ws.cell(row=row, column=c).number_format = '#,##0.00'


    def _set_title(ws, text, n_cols):
        col_letter = get_column_letter(n_cols)
        ws.merge_cells(f'A1:{col_letter}1')
        ws['A1'] = text
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = TITLE_FILL
        ws.row_dimensions[1].height = 22


    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


    # ─── Per-sheet appenders (one file at a time, appending into existing sheet) ──

    def _append_b2b(ws, data, src_name, alt_start):
        alt = alt_start
        for buyer in data.get('b2b', []):
            for inv in buyer.get('inv', []):
                for itm in inv.get('itms', []):
                    d = itm.get('itm_det', {})
                    ws.append([
                        src_name,
                        buyer['ctin'], buyer.get('cfs', ''),
                        inv.get('inum', ''), inv.get('idt', ''), inv.get('val', 0),
                        inv.get('pos', ''), inv.get('rchrg', ''), inv.get('inv_typ', ''),
                        itm.get('num', ''), d.get('rt', 0), d.get('txval', 0),
                        d.get('iamt', 0), d.get('camt', 0), d.get('samt', 0), d.get('csamt', 0),
                    ])
                    r = ws.max_row
                    _style_data(ws, r, 16, alt)
                    _fmt_num(ws, r, [6, 12, 13, 14, 15, 16])
                    alt = not alt
        return alt


    def _append_cdnr(ws, data, src_name, alt_start):
        alt = alt_start
        for rec in data.get('cdnr', []):
            for nt in rec.get('nt', []):
                for itm in nt.get('itms', []):
                    d = itm.get('itm_det', {})
                    ws.append([
                        src_name,
                        rec['ctin'], rec.get('cfs', ''),
                        nt.get('nt_num', ''), nt.get('nt_dt', ''), nt.get('ntty', ''),
                        nt.get('val', 0), nt.get('pos', ''), nt.get('rchrg', ''), nt.get('inv_typ', ''),
                        itm.get('num', ''), d.get('rt', 0), d.get('txval', 0),
                        d.get('iamt', 0), d.get('camt', 0), d.get('samt', 0), d.get('csamt', 0),
                    ])
                    r = ws.max_row
                    _style_data(ws, r, 17, alt)
                    _fmt_num(ws, r, [7, 13, 14, 15, 16, 17])
                    alt = not alt
        return alt


    def _append_exports(ws, data, src_name, alt_start):
        alt = alt_start
        for exp in data.get('exp', []):
            for inv in exp.get('inv', []):
                for itm in inv.get('itms', []):
                    ws.append([
                        src_name,
                        exp.get('exp_typ', ''),
                        inv.get('inum', ''), inv.get('idt', ''), inv.get('val', 0),
                        itm.get('rt', 0), itm.get('txval', 0),
                    ])
                    r = ws.max_row
                    _style_data(ws, r, 7, alt)
                    _fmt_num(ws, r, [5, 7])
                    alt = not alt
        return alt


    def _append_b2cs(ws, data, src_name, alt_start):
        alt = alt_start
        for rec in data.get('b2cs', []):
            ws.append([
                src_name,
                rec.get('typ', ''), rec.get('pos', ''), rec.get('sply_ty', ''), rec.get('rt', 0),
                rec.get('txval', 0), rec.get('iamt', 0), rec.get('camt', 0),
                rec.get('samt', 0), rec.get('csamt', 0),
            ])
            r = ws.max_row
            _style_data(ws, r, 10, alt)
            _fmt_num(ws, r, [5, 6, 7, 8, 9, 10])
            alt = not alt
        return alt


    def _append_hsn(ws, data, src_name, alt_start):
        alt = alt_start
        hsn_data = data.get('hsn', {})
        for cat, key in [('B2B', 'hsn_b2b'), ('B2C', 'hsn_b2c')]:
            for h in hsn_data.get(key, []):
                ws.append([
                    src_name,
                    cat, h.get('num', ''), h.get('hsn_sc', ''), h.get('desc', ''),
                    h.get('uqc', ''), h.get('qty', 0), h.get('rt', 0), h.get('txval', 0),
                    h.get('iamt', 0), h.get('camt', 0), h.get('samt', 0), h.get('csamt', 0),
                ])
                r = ws.max_row
                _style_data(ws, r, 13, alt)
                _fmt_num(ws, r, [7, 8, 9, 10, 11, 12, 13])
                alt = not alt
        return alt


    # ─── Workbook builder ─────────────────────────────────────────────────────────

    def build_consolidated_excel(file_records: list[dict], selected_sheets: dict) -> bytes:
        """
        file_records: [{'name': str, 'data': dict}, ...]
        selected_sheets: {sheet_name: bool}
        """
        wb = Workbook()
        wb.remove(wb.active)

        # ── Summary sheet ────────────────────────────────────────────────────────
        if selected_sheets.get('Summary'):
            ws = wb.create_sheet('Summary')
            hdr = ['Source File', 'GSTIN', 'Period', 'Filing Type', 'Filed On',
                   'GT', 'Current GT', 'B2B Buyers', 'B2B Invoices',
                   'CDNR Notes', 'Export Invoices', 'B2CS Records', 'HSN B2B Entries']
            n = len(hdr)
            ws.append([''] * n)
            ws.append(hdr)
            _style_hdr(ws, 2, n)
            for i, rec in enumerate(file_records):
                d = rec['data']
                gstin = d.get('gstin', '')
                fp    = d.get('fp', '')
                period = f"{fp[:2]}/{fp[2:]}" if len(fp) == 6 else fp
                row = [
                    rec['name'], gstin, period,
                    d.get('filing_typ', ''), d.get('fil_dt', ''),
                    d.get('gt', 0), d.get('cur_gt', 0),
                    len(d.get('b2b', [])),
                    sum(len(b['inv']) for b in d.get('b2b', [])),
                    sum(len(c['nt']) for c in d.get('cdnr', [])),
                    sum(len(e['inv']) for e in d.get('exp', [])),
                    len(d.get('b2cs', [])),
                    len(d.get('hsn', {}).get('hsn_b2b', [])),
                ]
                ws.append(row)
                _style_data(ws, ws.max_row, n, i % 2 == 0)
                _fmt_num(ws, ws.max_row, [6, 7])
            widths = [FILE_COL_W, 22, 10, 10, 12, 14, 14, 10, 10, 10, 12, 10, 12]
            _set_col_widths(ws, widths)
            _set_title(ws, 'GSTR-1 Consolidated Summary', n)
            ws.freeze_panes = 'A3'
            ws.row_dimensions[2].height = 30

        # ── B2B sheet ────────────────────────────────────────────────────────────
        if selected_sheets.get('B2B Invoices'):
            ws = wb.create_sheet('B2B Invoices')
            hdr = ['Source File', 'Buyer GSTIN', 'GSTIN Filing Status', 'Invoice No',
                   'Invoice Date', 'Invoice Value', 'Place of Supply', 'Reverse Charge',
                   'Invoice Type', 'Item No', 'Tax Rate (%)', 'Taxable Value',
                   'IGST', 'CGST', 'SGST', 'CESS']
            ws.append([''] * len(hdr))
            ws.append(hdr)
            _style_hdr(ws, 2, len(hdr))
            alt = False
            for rec in file_records:
                alt = _append_b2b(ws, rec['data'], rec['name'], alt)
            _set_col_widths(ws, [FILE_COL_W,20,8,18,12,14,8,8,8,8,10,14,14,14,14,10])
            _set_title(ws, 'B2B Invoices — Consolidated', len(hdr))
            ws.freeze_panes = 'A3'
            ws.row_dimensions[2].height = 30

        # ── CDNR sheet ───────────────────────────────────────────────────────────
        if selected_sheets.get('CDNR'):
            ws = wb.create_sheet('CDNR')
            hdr = ['Source File', 'Buyer GSTIN', 'GSTIN Filing Status', 'Note No',
                   'Note Date', 'Note Type', 'Note Value', 'Place of Supply',
                   'Reverse Charge', 'Invoice Type', 'Item No', 'Tax Rate (%)',
                   'Taxable Value', 'IGST', 'CGST', 'SGST', 'CESS']
            ws.append([''] * len(hdr))
            ws.append(hdr)
            _style_hdr(ws, 2, len(hdr))
            alt = False
            for rec in file_records:
                alt = _append_cdnr(ws, rec['data'], rec['name'], alt)
            _set_col_widths(ws, [FILE_COL_W,20,8,14,12,8,14,8,8,8,8,10,14,14,14,14,10])
            _set_title(ws, 'CDNR — Consolidated', len(hdr))
            ws.freeze_panes = 'A3'
            ws.row_dimensions[2].height = 30

        # ── Export Invoices sheet ────────────────────────────────────────────────
        if selected_sheets.get('Export Invoices'):
            ws = wb.create_sheet('Export Invoices')
            hdr = ['Source File', 'Export Type', 'Invoice No', 'Invoice Date',
                   'Invoice Value', 'Tax Rate (%)', 'Taxable Value']
            ws.append([''] * len(hdr))
            ws.append(hdr)
            _style_hdr(ws, 2, len(hdr))
            alt = False
            for rec in file_records:
                alt = _append_exports(ws, rec['data'], rec['name'], alt)
            _set_col_widths(ws, [FILE_COL_W, 12, 18, 12, 16, 10, 16])
            _set_title(ws, 'Export Invoices — Consolidated', len(hdr))
            ws.freeze_panes = 'A3'
            ws.row_dimensions[2].height = 30

        # ── B2CS sheet ───────────────────────────────────────────────────────────
        if selected_sheets.get('B2CS'):
            ws = wb.create_sheet('B2CS')
            hdr = ['Source File', 'Type', 'Place of Supply', 'Supply Type',
                   'Tax Rate (%)', 'Taxable Value', 'IGST', 'CGST', 'SGST', 'CESS']
            ws.append([''] * len(hdr))
            ws.append(hdr)
            _style_hdr(ws, 2, len(hdr))
            alt = False
            for rec in file_records:
                alt = _append_b2cs(ws, rec['data'], rec['name'], alt)
            _set_col_widths(ws, [FILE_COL_W, 10, 8, 12, 10, 16, 14, 14, 14, 10])
            _set_title(ws, 'B2CS — Consolidated', len(hdr))
            ws.freeze_panes = 'A3'
            ws.row_dimensions[2].height = 30

        # ── HSN sheet ────────────────────────────────────────────────────────────
        if selected_sheets.get('HSN Summary'):
            ws = wb.create_sheet('HSN Summary')
            hdr = ['Source File', 'Category', 'S.No', 'HSN Code', 'Description',
                   'UQC', 'Quantity', 'Tax Rate (%)', 'Taxable Value',
                   'IGST', 'CGST', 'SGST', 'CESS']
            ws.append([''] * len(hdr))
            ws.append(hdr)
            _style_hdr(ws, 2, len(hdr))
            alt = False
            for rec in file_records:
                alt = _append_hsn(ws, rec['data'], rec['name'], alt)
            _set_col_widths(ws, [FILE_COL_W, 10, 6, 12, 36, 8, 10, 10, 16, 14, 14, 14, 10])
            _set_title(ws, 'HSN Summary — Consolidated', len(hdr))
            ws.freeze_panes = 'A3'
            ws.row_dimensions[2].height = 30

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    # ─── Helper: parse uploaded file (JSON or ZIP) → list of {name, data} ─────────

    def parse_upload(uploaded_file) -> list[dict]:
        raw   = uploaded_file.read()
        fname = uploaded_file.name          # this is always the display name
        records = []

        if fname.lower().endswith('.zip'):
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                json_names = [n for n in zf.namelist() if n.endswith('.json') and not n.startswith('__MACOSX')]
                for jname in json_names:
                    try:
                        data = json.loads(zf.read(jname))
                        # ← use the ZIP filename, not the inner JSON name
                        records.append({'name': fname, 'data': data})
                    except Exception:
                        pass
        else:
            try:
                data = json.loads(raw)
                records.append({'name': fname, 'data': data})
            except Exception:
                pass

        return records


    # ─── Streamlit UI ─────────────────────────────────────────────────────────────


    st.title('🧾 GSTR-1 JSON → Excel Converter')
    st.markdown(
        'Upload **one or more** GSTR-1 JSON / ZIP files. '
        'All filings will be consolidated into a single Excel workbook — '
        'each section as its own sheet with a **Source File** column identifying each transaction.'
    )

    uploaded_files = st.file_uploader(
        'Upload GSTR-1 JSON or ZIP files (multiple allowed)',
        type=['json', 'zip'],
        accept_multiple_files=True,
    )

    if uploaded_files:
        # ── Parse all files ──────────────────────────────────────────────────────
        all_records: list[dict] = []
        parse_errors = []

        for uf in uploaded_files:
            recs = parse_upload(uf)
            if recs:
                all_records.extend(recs)
            else:
                parse_errors.append(uf.name)

        if parse_errors:
            st.warning(f"⚠️ Could not parse: {', '.join(parse_errors)}")

        if not all_records:
            st.error('No valid GSTR-1 JSON data found in the uploaded files.')
            st.stop()

        # ── Per-file summary table ───────────────────────────────────────────────
        st.divider()
        st.subheader(f'📂 {len(all_records)} file(s) loaded')

        tbl_rows = []
        for rec in all_records:
            d = rec['data']
            fp = d.get('fp', '')
            tbl_rows.append({
                'Source File':      rec['name'],
                'GSTIN':            d.get('gstin', '-'),
                'Period':           f"{fp[:2]}/{fp[2:]}" if len(fp) == 6 else fp,
                'Filing Type':      d.get('filing_typ', '-'),
                'Filed On':         d.get('fil_dt', '-'),
                'B2B Invoices':     sum(len(b['inv']) for b in d.get('b2b', [])),
                'CDNR Notes':       sum(len(c['nt']) for c in d.get('cdnr', [])),
                'Export Invoices':  sum(len(e['inv']) for e in d.get('exp', [])),
                'B2CS Records':     len(d.get('b2cs', [])),
            })

        st.dataframe(tbl_rows, use_container_width=True, hide_index=True)

        # ── Totals row ───────────────────────────────────────────────────────────
        c1, c2, c3, c4 = st.columns(4)
        c1.metric('Total B2B Invoices',    sum(r['B2B Invoices']    for r in tbl_rows))
        c2.metric('Total CDNR Notes',      sum(r['CDNR Notes']      for r in tbl_rows))
        c3.metric('Total Export Invoices', sum(r['Export Invoices'] for r in tbl_rows))
        c4.metric('Total B2CS Records',    sum(r['B2CS Records']    for r in tbl_rows))

        # ── Sheet selector ───────────────────────────────────────────────────────
        st.divider()
        st.subheader('⚙️ Select Sheets to Include')

        has = {
            'Summary':         True,
            'B2B Invoices':    any(rec['data'].get('b2b')  for rec in all_records),
            'CDNR':            any(rec['data'].get('cdnr') for rec in all_records),
            'Export Invoices': any(rec['data'].get('exp')  for rec in all_records),
            'B2CS':            any(rec['data'].get('b2cs') for rec in all_records),
            'HSN Summary':     any(
                rec['data'].get('hsn', {}).get('hsn_b2b') or rec['data'].get('hsn', {}).get('hsn_b2c')
                for rec in all_records
            ),
        }
        selected = {}
        cols = st.columns(3)
        for i, (sheet, avail) in enumerate(has.items()):
            label = sheet if avail else f'{sheet} (no data)'
            selected[sheet] = cols[i % 3].checkbox(label, value=avail, disabled=not avail)

        # ── Generate button ──────────────────────────────────────────────────────
        st.divider()
        if st.button('🚀 Generate Consolidated Excel', use_container_width=True, type='primary'):
            if not any(selected.values()):
                st.warning('Please select at least one sheet.')
                st.stop()

            with st.spinner('Building consolidated Excel workbook…'):
                excel_bytes = build_consolidated_excel(all_records, selected)

            # derive output filename from date range of filings
            periods = sorted({rec['data'].get('fp','') for rec in all_records if rec['data'].get('fp')})
            period_tag = f"{periods[0]}_to_{periods[-1]}" if len(periods) > 1 else (periods[0] if periods else 'multi')
            out_name = f'GSTR1_Consolidated_{period_tag}.xlsx'

            st.success(f'✅ Consolidated workbook ready — **{out_name}**  ({len(all_records)} filing(s), {sum(selected.values())} sheet(s))')
            st.download_button(
                label='⬇️ Download Excel',
                data=excel_bytes,
                file_name=out_name,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )

    st.divider()
    st.caption(
        'Supports: multiple JSON files · ZIP files containing multiple JSONs · '
        'Sections: Summary · B2B · CDNR · Exports · B2CS · HSN  |  '
        'Each row tagged with its Source File for full traceability'
    )


# ── GSTR-2A Combiner ────────────────────────────────────────────
elif ACTIVE == "GSTR-2A Combiner":
    page_header("📊", "GSTR-2A File Combiner", "Upload multiple GSTR-2A .xlsx files to combine them sheet-wise, removing subtotal rows (Rate = '-').")


    st.title("📊 GSTR-2A File Combiner")
    st.markdown("Upload multiple GSTR-2A `.xlsx` files to combine them sheet-wise, removing subtotal rows (rows where **Rate = `-`**).")

    # ── Sheet definitions ─────────────────────────────────────────────────────────
    # Each sheet has: header_row (0-indexed), data_start_row, rate_col_index (0-indexed, None if no Rate col)
    SHEET_CONFIG = {
        "B2B":      {"header_rows": [4, 5], "data_start": 6, "rate_col": 8},
        "B2BA":     {"header_rows": [5, 6], "data_start": 7, "rate_col": 10},
        "CDNR":     {"header_rows": [4, 5], "data_start": 6, "rate_col": 9},
        "CDNRA":    {"header_rows": [5, 6], "data_start": 7, "rate_col": 12},
        "ECO":      {"header_rows": [4, 5], "data_start": 6, "rate_col": 7},
        "ECOA":     {"header_rows": [5, 6], "data_start": 7, "rate_col": 9},
        "ISD":      {"header_rows": [4, 5], "data_start": 6, "rate_col": None},
        "ISDA":     {"header_rows": [5, 6], "data_start": 7, "rate_col": None},
        "TDS":      {"header_rows": [4, 5], "data_start": 6, "rate_col": None},
        "TDSA":     {"header_rows": [4, 5], "data_start": 6, "rate_col": None},
        "TCS":      {"header_rows": [4, 5], "data_start": 6, "rate_col": None},
        "IMPG":     {"header_rows": [4, 5], "data_start": 6, "rate_col": None},
        "IMPG SEZ": {"header_rows": [4, 5], "data_start": 6, "rate_col": None},
    }

    MONTH_NAMES = {
        "01": "January",   "02": "February", "03": "March",
        "04": "April",     "05": "May",       "06": "June",
        "07": "July",      "08": "August",    "09": "September",
        "10": "October",   "11": "November",  "12": "December",
    }

    def parse_date_from_filename(filename: str) -> str:
        """
        Extract date label from GSTR-2A filename.
        Pattern: <GSTIN>_MMYYYY_R2A.xlsx  e.g. 29AACCZ2032J1Z8_022026_R2A.xlsx
        Returns e.g. "February 2026"
        """
        # Match 6-digit segment that looks like MMYYYY
        match = re.search(r'_(\d{2})(\d{4})_', filename)
        if match:
            mm, yyyy = match.group(1), match.group(2)
            month_name = MONTH_NAMES.get(mm, mm)
            return f"{month_name} {yyyy}"
        return "Unknown"


    def flatten_headers(row1: list, row2: list) -> list:
        """
        Merge two header rows into one flat row.
        row2 wins if it has a value (specific sub-header).
        row1 wins if row2 is NaN (standalone header with no sub-columns).
        """
        result = []
        for r1, r2 in zip(row1, row2):
            r1_val = str(r1).strip() if pd.notna(r1) and str(r1).strip() not in ("", "nan") else None
            r2_val = str(r2).strip() if pd.notna(r2) and str(r2).strip() not in ("", "nan") else None
            if r2_val:
                result.append(r2_val)
            elif r1_val:
                result.append(r1_val)
            else:
                result.append("")
        return result


    def extract_sheet_data(xl_file, sheet_name, file_date: str):
        """
        Read a sheet, return:
          - header_rows: single-element list with one flat merged header row
          - data_df: DataFrame of actual data rows, with subtotals removed
                     and a 'Period' column inserted as the first column
          - removed_count: how many subtotal rows were dropped
        """
        cfg = SHEET_CONFIG.get(sheet_name)
        if cfg is None:
            return None, None, 0

        df_raw = pd.read_excel(xl_file, sheet_name=sheet_name, header=None)

        # Flatten the two header rows into one
        r_idx = cfg["header_rows"]
        row1 = df_raw.iloc[r_idx[0]].tolist()
        row2 = df_raw.iloc[r_idx[1]].tolist()
        flat = flatten_headers(row1, row2)
        # Prepend Period column label
        header_rows = [["Period"] + flat]

        # Extract data rows
        data = df_raw.iloc[cfg["data_start"]:].reset_index(drop=True)

        # Drop rows that are entirely NaN (blank spacer rows between supplier groups)
        data = data.dropna(how="all").reset_index(drop=True)

        removed_count = 0
        rate_col = cfg["rate_col"]
        if rate_col is not None and rate_col < data.shape[1]:
            # Subtotal rows: Rate column == '-' (string dash, may have spaces)
            mask = data.iloc[:, rate_col].astype(str).str.strip() == "-"
            removed_count = mask.sum()
            data = data[~mask].reset_index(drop=True)

        # Insert Period as the leftmost column
        data.insert(0, "Period", file_date)

        return header_rows, data, removed_count


    def write_combined_excel(combined: dict, sheet_configs: dict) -> bytes:
        """
        Write the combined data to an Excel workbook.
        Returns bytes of the .xlsx file.
        """
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        header_font = Font(name="Arial", bold=True, size=10)
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        data_font = Font(name="Arial", size=10)
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for sheet_name, (header_rows, data_df) in combined.items():
            ws = wb.create_sheet(title=sheet_name)

            if header_rows is None or data_df is None:
                continue

            # Period column gets a distinct highlight color
            period_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow

            # Write single flat header row
            for hr in header_rows:
                ws.append([str(v) if v else "" for v in hr])
                for cell in ws[ws.max_row]:
                    cell.font = header_font
                    cell.fill = period_fill if cell.column == 1 else header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    cell.border = border
            # Freeze the header row
            ws.freeze_panes = "A2"

            # Write data rows
            for _, row in data_df.iterrows():
                values = []
                for v in row:
                    if pd.isna(v):
                        values.append("")
                    else:
                        values.append(v)
                ws.append(values)
                for cell in ws[ws.max_row]:
                    cell.font = data_font
                    cell.border = border

            # Auto-size columns (capped at 50)
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            max_len = max(max_len, len(str(cell.value or "")))
                        except Exception:
                            pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()


    # ── UI ────────────────────────────────────────────────────────────────────────
    uploaded_files = st.file_uploader(
        "Upload GSTR-2A Excel files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="You can upload 2 or more GSTR-2A files at once."
    )

    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} file(s) uploaded.")

        # Show file details
        with st.expander("📁 Uploaded files", expanded=True):
            for f in uploaded_files:
                st.write(f"• **{f.name}** ({f.size / 1024:.1f} KB)")

        if st.button("🔄 Combine Files", type="primary"):
            combined = {}       # sheet_name -> (header_rows, combined_df)
            stats = {}          # sheet_name -> {files: n, rows: n, removed: n}

            progress = st.progress(0, text="Processing files…")
            all_sheets = list(SHEET_CONFIG.keys())

            for s_idx, sheet_name in enumerate(all_sheets):
                progress.progress((s_idx) / len(all_sheets), text=f"Processing sheet: **{sheet_name}**")

                all_data_frames = []
                saved_headers = None
                total_removed = 0
                files_with_data = 0

                for uf in uploaded_files:
                    uf.seek(0)
                    file_date = parse_date_from_filename(uf.name)
                    try:
                        xl = pd.ExcelFile(uf)
                        if sheet_name not in xl.sheet_names:
                            continue
                        header_rows, data_df, removed = extract_sheet_data(uf, sheet_name, file_date)
                        if data_df is not None and len(data_df) > 0:
                            all_data_frames.append(data_df)
                            files_with_data += 1
                        if saved_headers is None and header_rows is not None:
                            saved_headers = header_rows
                        total_removed += removed
                    except Exception as e:
                        st.warning(f"⚠️ Could not read sheet **{sheet_name}** from `{uf.name}`: {e}")

                if all_data_frames:
                    combined_df = pd.concat(all_data_frames, ignore_index=True)
                else:
                    combined_df = pd.DataFrame()

                combined[sheet_name] = (saved_headers, combined_df)
                stats[sheet_name] = {
                    "files": files_with_data,
                    "rows": len(combined_df),
                    "removed": total_removed,
                }

            progress.progress(1.0, text="✅ All sheets processed!")

            # Summary table
            st.subheader("📋 Combination Summary")
            summary_rows = []
            for sn, s in stats.items():
                summary_rows.append({
                    "Sheet": sn,
                    "Files with data": s["files"],
                    "Data rows combined": s["rows"],
                    "Subtotal rows removed": s["removed"],
                })
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

            # Generate output Excel
            with st.spinner("Generating combined Excel file…"):
                excel_bytes = write_combined_excel(combined, SHEET_CONFIG)

            st.download_button(
                label="⬇️ Download Combined GSTR-2A Excel",
                data=excel_bytes,
                file_name="GSTR2A_Combined.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # Preview section
            st.subheader("🔍 Preview Combined Data")
            preview_sheet = st.selectbox(
                "Select sheet to preview",
                [s for s in all_sheets if stats[s]["rows"] > 0]
            )
            if preview_sheet and combined[preview_sheet][1] is not None and len(combined[preview_sheet][1]) > 0:
                hdr, df = combined[preview_sheet]
                # hdr is now a single-element list with the flat header row
                if hdr and len(hdr) == 1:
                    col_names = [str(v) if v else f"Col{i}" for i, v in enumerate(hdr[0])]
                    if len(col_names) == df.shape[1]:
                        df = df.copy()
                        df.columns = col_names
                st.dataframe(df.head(50), use_container_width=True, hide_index=True)
                if len(df) > 50:
                    st.caption(f"Showing first 50 of {len(df)} rows.")

    else:
        st.info("👆 Please upload one or more GSTR-2A Excel files to get started.")

    st.markdown("---")
    st.caption("GSTR-2A Combiner • Removes subtotal rows (Rate = `-`) • Combines sheet-wise across all uploaded files")


# ── GSTR-2B Consolidator ────────────────────────────────────────
elif ACTIVE == "GSTR-2B Consolidator":
    page_header("📋", "GSTR-2B Consolidator", "Upload multiple GSTR-2B Excel files to merge them into one clean consolidated workbook, sheet-wise.")


    st.markdown("""
    <style>
        .main-header {
            background: linear-gradient(135deg, #1a237e 0%, #283593 50%, #3949ab 100%);
            padding: 2rem 2.5rem; border-radius: 12px; margin-bottom: 2rem; color: white;
        }
        .main-header h1 { color: white; margin: 0; font-size: 2.2rem; font-weight: 700; }
        .main-header p  { color: #c5cae9; margin: 0.4rem 0 0; font-size: 1rem; }
        .stat-card {
            background: white; border: 1px solid #e3e8f0; border-radius: 10px;
            padding: 1.2rem 1.5rem; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }
        .stat-card .value { font-size: 2rem; font-weight: 700; color: #1a237e; }
        .stat-card .label { font-size: 0.85rem; color: #666; margin-top: 0.2rem; }
        .success-box {
            background: #e8f5e9; border-left: 4px solid #43a047;
            border-radius: 6px; padding: 1rem 1.2rem; margin: 1rem 0;
        }
        .sheet-pill {
            display: inline-block; background: #e8eaf6; color: #3949ab;
            border-radius: 20px; padding: 0.2rem 0.8rem; font-size: 0.8rem;
            margin: 0.2rem; font-weight: 500;
        }
        .stDownloadButton > button {
            background: linear-gradient(135deg, #1a237e, #3949ab);
            color: white; border: none; border-radius: 8px;
            padding: 0.7rem 2rem; font-size: 1rem; font-weight: 600;
            width: 100%;
        }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="main-header">
      <h1>📊 GSTR-2B Consolidator</h1>
      <p>Upload multiple GSTR-2B Excel files to merge them into one clean consolidated workbook, sheet-wise</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Constants ─────────────────────────────────────────────────────────────────

    # Sheets to SKIP (summary/narrative sheets, not transaction data)
    SKIP_SHEETS = {"ITC Available", "ITC not available", "ITC Reversal", "ITC Rejected"}

    # Sheets that have THREE header rows (rows 4, 5, 6) instead of two (rows 4, 5)
    # These are "amendment" sheets with "Original Details / Revised Details" split header
    AMENDMENT_SHEETS = {
        "B2BA", "B2B-CDNRA", "B2B-DNRA",
        "B2BA (ITC Reversal)", "ECOA", "ISDA",
        "B2BA(Rejected)", "B2B-CDNRA(Rejected)", "ECOA(Rejected)", "ISDA(Rejected)",
    }

    SHEET_DESC = {
        "Read me":              "Return metadata – GSTIN, period, generation date",
        "B2B":                  "Taxable inward supplies from registered persons",
        "B2BA":                 "Amendments to B2B invoices",
        "B2B-CDNR":             "Credit / Debit notes from registered persons",
        "B2B-CDNRA":            "Amendments to Credit / Debit notes",
        "ECO":                  "Documents via E-Commerce Operator (u/s 9(5))",
        "ECOA":                 "Amendments to ECO documents",
        "ISD":                  "Input Service Distributor credits",
        "ISDA":                 "Amendments to ISD credits",
        "IMPG":                 "Imports from overseas (Bill of Entry)",
        "IMPGA":                "Amendments – overseas imports",
        "IMPGSEZ":              "Imports from SEZ units (Bill of Entry)",
        "IMPGSEZA":             "Amendments – SEZ imports",
        "B2B (ITC Reversal)":   "B2B invoices attracting ITC reversal",
        "B2BA (ITC Reversal)":  "Amendments – B2B ITC reversal",
        "B2B-DNR":              "Debit notes (original)",
        "B2B-DNRA":             "Amendments to Debit notes",
        "B2B(Rejected)":        "B2B invoices – ITC rejected",
        "B2BA(Rejected)":       "B2BA invoices – ITC rejected",
        "B2B-CDNR(Rejected)":   "Credit / Debit notes – ITC rejected",
        "B2B-CDNRA(Rejected)":  "CDN amendments – ITC rejected",
        "ECO(Rejected)":        "ECO documents – ITC rejected",
        "ECOA(Rejected)":       "ECOA documents – ITC rejected",
        "ISD(Rejected)":        "ISD credits – ITC rejected",
        "ISDA(Rejected)":       "ISDA credits – ITC rejected",
    }

    # ── Excel styles ──────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1A237E")
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    META_FILL = PatternFill("solid", fgColor="E8EAF6")
    DATA_FONT = Font(name="Arial", size=9)
    ALT_FILL  = PatternFill("solid", fgColor="F5F6FF")
    WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
    THIN      = Side(style="thin", color="C5CAE9")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Helpers ───────────────────────────────────────────────────────────────────

    def sort_key(fname: str) -> int:
        try:
            return int(os.path.basename(fname)[:6])
        except Exception:
            return 999999


    def get_period_label(fname: str) -> str:
        prefix = os.path.basename(fname)[:6]
        try:
            mm, yy = int(prefix[:2]), int(prefix[2:])
            months = ["Jan","Feb","Mar","Apr","May","Jun",
                      "Jul","Aug","Sep","Oct","Nov","Dec"]
            return f"{months[mm-1]}-{yy}"
        except Exception:
            return prefix


    def clean_headers(raw: list[str]) -> list[str]:
        """Deduplicate blank/repeated column names."""
        seen: dict[str, int] = {}
        result = []
        for h in raw:
            h = h.strip()
            if not h:
                h = "Column"
            if h in seen:
                seen[h] += 1
                result.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                result.append(h)
        return result


    def parse_sheet(xl: pd.ExcelFile, sheet: str) -> tuple[list[str], pd.DataFrame]:
        """
        Parse a GSTR-2B sheet and return (headers, data_df).

        Normal sheets   : rows 0-3 = junk/title, row 4 = main headers,
                          row 5 = sub-headers, rows 6+ = data.
        Amendment sheets: rows 0-3 = junk/title, row 4 = group labels (Original/Revised),
                          row 5 = main headers, row 6 = sub-headers, rows 7+ = data.
        """
        raw = pd.read_excel(xl, sheet_name=sheet, header=None, dtype=str).fillna("")
        n   = len(raw)

        is_amendment = sheet in AMENDMENT_SHEETS

        if is_amendment:
            # Need at least 7 rows to have any data
            if n < 7:
                hdrs = _build_amendment_headers(raw) if n >= 7 else ["Column"]
                return hdrs, pd.DataFrame(columns=hdrs)
            headers  = _build_amendment_headers(raw)
            data_raw = raw.iloc[7:].copy()
        else:
            if n < 6:
                hdrs = _build_normal_headers(raw) if n >= 6 else ["Column"]
                return hdrs, pd.DataFrame(columns=hdrs)
            headers  = _build_normal_headers(raw)
            data_raw = raw.iloc[6:].copy()

        # Align columns
        data_raw = data_raw.reset_index(drop=True)
        data_raw.columns = range(len(data_raw.columns))
        ncols = len(headers)
        # Pad if fewer columns
        for i in range(ncols - len(data_raw.columns)):
            data_raw[len(data_raw.columns)] = ""
        data_raw = data_raw.iloc[:, :ncols]
        data_raw.columns = headers

        # Drop fully-empty rows
        data_raw = data_raw[~(data_raw == "").all(axis=1)].reset_index(drop=True)
        return headers, data_raw


    def _build_normal_headers(raw: pd.DataFrame) -> list[str]:
        """Combine rows 4 and 5 into one header list."""
        row4 = [str(v).strip() for v in raw.iloc[4]] if len(raw) > 4 else []
        row5 = [str(v).strip() for v in raw.iloc[5]] if len(raw) > 5 else []
        ncols = max(len(row4), len(row5))
        row4 += [""] * (ncols - len(row4))
        row5 += [""] * (ncols - len(row5))
        combined = []
        for a, b in zip(row4, row5):
            if a and b and a != b:
                combined.append(f"{a} - {b}")
            elif a:
                combined.append(a)
            elif b:
                combined.append(b)
            else:
                combined.append("")
        return clean_headers(combined)


    def _build_amendment_headers(raw: pd.DataFrame) -> list[str]:
        """
        Amendment sheets have 3 header rows:
          row4 = group labels  (Original Details | Revised Details | …)
          row5 = main headers
          row6 = sub-headers
        Strategy: combine all three rows, propagating group labels forward.
        """
        row4 = [str(v).strip() for v in raw.iloc[4]] if len(raw) > 4 else []
        row5 = [str(v).strip() for v in raw.iloc[5]] if len(raw) > 5 else []
        row6 = [str(v).strip() for v in raw.iloc[6]] if len(raw) > 6 else []

        ncols = max(len(row4), len(row5), len(row6))
        row4 += [""] * (ncols - len(row4))
        row5 += [""] * (ncols - len(row5))
        row6 += [""] * (ncols - len(row6))

        # Forward-fill group label in row4
        current_group = ""
        row4_filled = []
        for v in row4:
            if v:
                current_group = v
            row4_filled.append(current_group)

        combined = []
        for g, a, b in zip(row4_filled, row5, row6):
            parts = []
            # Only prefix group if it adds context (not empty, not same as a)
            if g and g not in (a, b) and g not in ("Original Details", "Revised Details",
                                                      "Whether ITC to be reduced (Taxpayer's Input)",
                                                      "Amount declared by taxpayer for ITC reduction"):
                pass  # skip generic group labels – they clutter headers
            if a and b and a != b:
                parts.append(f"{a} - {b}")
            elif a:
                parts.append(a)
            elif b:
                parts.append(b)
            combined.append(" ".join(parts) if parts else "")

        return clean_headers(combined)


    def col_width(series: pd.Series, header: str) -> float:
        max_len = max(
            len(str(header)),
            series.astype(str).str.len().max() if len(series) else 0
        )
        return min(max(max_len * 1.15, 10), 50)


    # ── Excel writer ──────────────────────────────────────────────────────────────

    def write_consolidated_excel(
        all_data: dict,
        period_labels: list[str],
        sheet_names: list[str],
        period_meta: list[dict],      # [{period, gstin, legal_name, tax_period, fin_year, gen_date}]
    ) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        # ── Read me sheet ─────────────────────────────────────────────────────────
        ws_rm = wb.create_sheet("Read me")
        ws_rm.sheet_view.showGridLines = False
        ws_rm.column_dimensions["A"].width = 28
        ws_rm.column_dimensions["B"].width = 35

        # Title
        ws_rm.merge_cells("A1:B1")
        ws_rm["A1"] = "GSTR-2B CONSOLIDATED WORKBOOK"
        ws_rm["A1"].font  = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        ws_rm["A1"].fill  = PatternFill("solid", fgColor="1A237E")
        ws_rm["A1"].alignment = CENTER
        ws_rm.row_dimensions[1].height = 28

        # Common entity info from first period
        first = period_meta[0] if period_meta else {}
        common_rows = [
            ("GSTIN",        first.get("gstin", "")),
            ("Legal Name",   first.get("legal_name", "")),
            ("Financial Year", first.get("fin_year", "")),
            ("Consolidated Periods", ", ".join(p["tax_period"] for p in period_meta)),
        ]
        for r, (k, v) in enumerate(common_rows, start=2):
            c1 = ws_rm.cell(r, 1, k)
            c1.font  = Font(name="Arial", bold=True, size=10)
            c1.fill  = META_FILL
            c1.border = BORDER
            c2 = ws_rm.cell(r, 2, v)
            c2.font  = Font(name="Arial", size=10)
            c2.border = BORDER
            ws_rm.row_dimensions[r].height = 18

        # Per-period tax period listing
        sep_row = len(common_rows) + 2 + 1
        ws_rm.cell(sep_row, 1, "Period Details").font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        ws_rm.cell(sep_row, 1).fill = PatternFill("solid", fgColor="283593")
        ws_rm.cell(sep_row, 1).border = BORDER
        ws_rm.cell(sep_row, 2, "Generation Date").font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        ws_rm.cell(sep_row, 2).fill = PatternFill("solid", fgColor="283593")
        ws_rm.cell(sep_row, 2).border = BORDER
        ws_rm.row_dimensions[sep_row].height = 18

        for i, pm in enumerate(period_meta, start=sep_row + 1):
            fill = ALT_FILL if i % 2 == 0 else WHT_FILL
            ws_rm.cell(i, 1, pm.get("tax_period", "")).font   = Font(name="Arial", size=10)
            ws_rm.cell(i, 1).fill   = fill
            ws_rm.cell(i, 1).border = BORDER
            ws_rm.cell(i, 2, pm.get("gen_date", "")).font   = Font(name="Arial", size=10)
            ws_rm.cell(i, 2).fill   = fill
            ws_rm.cell(i, 2).border = BORDER
            ws_rm.row_dimensions[i].height = 18

        # ── Data sheets ───────────────────────────────────────────────────────────
        for sheet_name in sheet_names:
            frames = []
            for plab in period_labels:
                df = all_data.get(plab, {}).get(sheet_name, pd.DataFrame())
                if not df.empty:
                    df = df.copy()
                    df.insert(0, "Tax Period", plab)
                    frames.append(df)

            safe_name = sheet_name[:31]
            ws = wb.create_sheet(safe_name)
            ws.sheet_view.showGridLines = False

            if not frames:
                ws["A1"] = "No data available for this section across the selected periods."
                ws["A1"].font = Font(name="Arial", italic=True, color="888888", size=9)
                continue

            combined = pd.concat(frames, ignore_index=True, sort=False).fillna("")

            ws.freeze_panes = "B3"

            # Title row
            title_text = SHEET_DESC.get(sheet_name, sheet_name)
            ncols = max(len(combined.columns), 2)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            ws["A1"] = f"GSTR-2B  |  {sheet_name}  –  {title_text}"
            ws["A1"].font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            ws["A1"].fill      = PatternFill("solid", fgColor="1A237E")
            ws["A1"].alignment = CENTER
            ws.row_dimensions[1].height = 24

            # Header row
            for c, col in enumerate(combined.columns, start=1):
                cell = ws.cell(row=2, column=c, value=col)
                cell.font      = HDR_FONT
                cell.fill      = HDR_FILL
                cell.alignment = CENTER
                cell.border    = BORDER
            ws.row_dimensions[2].height = 30

            # Data rows
            for r_idx, row in enumerate(combined.itertuples(index=False), start=3):
                fill = ALT_FILL if r_idx % 2 == 0 else WHT_FILL
                for c_idx, val in enumerate(row, start=1):
                    v = "" if str(val) in ("nan", "None") else val
                    cell = ws.cell(row=r_idx, column=c_idx, value=v)
                    cell.font      = DATA_FONT
                    cell.fill      = fill
                    cell.border    = BORDER
                    cell.alignment = LEFT
                ws.row_dimensions[r_idx].height = 15

            # Column widths
            for c_idx, col in enumerate(combined.columns, start=1):
                ws.column_dimensions[get_column_letter(c_idx)].width = col_width(combined[col], col)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    # ── Streamlit UI ──────────────────────────────────────────────────────────────

    uploaded = st.file_uploader(
        "Upload GSTR-2B Excel Files (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Upload one or more GSTR-2B Excel files. Files are auto-sorted by period (MMYYYY prefix)."
    )

    if not uploaded:
        st.info("⬆️  Upload at least one GSTR-2B file to get started.")
        st.stop()

    uploaded_sorted = sorted(uploaded, key=lambda f: sort_key(f.name))

    with st.spinner("Reading and parsing files…"):
        all_data:      dict[str, dict[str, pd.DataFrame]] = {}
        period_labels: list[str]  = []
        period_meta:   list[dict] = []
        sheet_names:   list[str]  = []
        errors:        list[str]  = []

        for uf in uploaded_sorted:
            try:
                xl     = pd.ExcelFile(uf)
                period = get_period_label(uf.name)
                period_labels.append(period)
                all_data[period] = {}

                # Determine sheets to consolidate (from first file, minus skipped)
                if not sheet_names:
                    sheet_names = [s for s in xl.sheet_names
                                   if s not in SKIP_SHEETS and s != "Read me"]

                # Parse each sheet
                for sname in xl.sheet_names:
                    if sname in SKIP_SHEETS or sname == "Read me":
                        continue
                    try:
                        _, df = parse_sheet(xl, sname)
                        all_data[period][sname] = df
                    except Exception as e:
                        errors.append(f"{uf.name} / {sname}: {e}")

                # Extract metadata from Read me
                meta = {"period": period, "gstin": "", "legal_name": "",
                        "tax_period": "", "fin_year": "", "gen_date": ""}
                try:
                    rm = pd.read_excel(xl, sheet_name="Read me", header=None, dtype=str).fillna("")
                    for _, row in rm.iterrows():
                        key = str(row.iloc[0]).strip()
                        val = str(row.iloc[2]).strip() if len(row) > 2 else ""
                        if key == "GSTIN":           meta["gstin"]       = val
                        if key == "Legal Name":      meta["legal_name"]  = val
                        if key == "Tax Period":      meta["tax_period"]  = val
                        if key == "Financial Year":  meta["fin_year"]    = val
                        if key == "Date of generation": meta["gen_date"] = val
                except Exception:
                    meta["tax_period"] = period
                period_meta.append(meta)

            except Exception as e:
                errors.append(f"{uf.name}: {e}")

    if errors:
        with st.expander("⚠️  Warnings during parsing"):
            for e in errors:
                st.warning(e)

    # ── Stats ──────────────────────────────────────────────────────────────────────
    total_rows = sum(
        df.shape[0]
        for pdata in all_data.values()
        for df in pdata.values()
    )
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="stat-card"><div class="value">{len(uploaded)}</div>'
                    f'<div class="label">Files Uploaded</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-card"><div class="value">{len(period_labels)}</div>'
                    f'<div class="label">Periods</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-card"><div class="value">{len(sheet_names)}</div>'
                    f'<div class="label">Sheets Consolidated</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="stat-card"><div class="value">{total_rows:,}</div>'
                    f'<div class="label">Total Data Rows</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**📁 Files Detected**")
        for i, uf in enumerate(uploaded_sorted):
            plab = period_labels[i] if i < len(period_labels) else "?"
            tp   = period_meta[i]["tax_period"] if i < len(period_meta) else ""
            st.markdown(f"- `{uf.name}` → **{tp} ({plab})**")
    with col_b:
        first_meta = period_meta[0] if period_meta else {}
        st.markdown("**🏢 Entity Details**")
        st.markdown(f"**GSTIN:** `{first_meta.get('gstin', 'Not found')}`")
        st.markdown(f"**Legal Name:** `{first_meta.get('legal_name', 'Not found')}`")
        st.markdown(f"**Financial Year:** `{first_meta.get('fin_year', 'Not found')}`")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Sheet preview ──────────────────────────────────────────────────────────────
    with st.expander("🔍 Preview Sheet Data"):
        nonempty = [s for s in sheet_names if any(
            not all_data.get(p, {}).get(s, pd.DataFrame()).empty for p in period_labels
        )]
        if nonempty:
            sel_sheet  = st.selectbox("Select Sheet",  nonempty)
            sel_period = st.selectbox("Select Period", period_labels)
            preview    = all_data.get(sel_period, {}).get(sel_sheet, pd.DataFrame())
            if preview.empty:
                st.info("No data for this sheet/period combination.")
            else:
                st.dataframe(preview.head(50), use_container_width=True)
                st.caption(f"Showing up to 50 of {len(preview)} rows")

    st.markdown("---")
    st.subheader("📥 Download Consolidated Report")

    with st.spinner("Building Excel workbook…"):
        excel_bytes = write_consolidated_excel(
            all_data=all_data,
            period_labels=period_labels,
            sheet_names=sheet_names,
            period_meta=period_meta,
        )

    first_meta = period_meta[0] if period_meta else {}
    gstin      = first_meta.get("gstin", "GSTIN")
    periods_str = "_".join(p["tax_period"] for p in period_meta) or "_".join(period_labels)
    filename   = "GSTR-2B Consolidated.xlsx"

    st.markdown(f"""
    <div class="success-box">
    ✅ Workbook ready — <strong>{len(sheet_names) + 1} sheets</strong>
    (1 Read me + {len(sheet_names)} transaction sections across {len(period_labels)} period(s))<br>
    <strong>Note:</strong> ITC Available, ITC Not Available, ITC Reversal and ITC Rejected summary sheets are excluded (not transaction data).
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        label="⬇️  Download GSTR-2B Consolidated Excel",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("**Sheets in the workbook:**")
    pills = " ".join(f'<span class="sheet-pill">{s}</span>'
                     for s in ["Read me"] + sheet_names)
    st.markdown(pills, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.caption("Sheets excluded (summary, not transaction data): " +
               " | ".join(SKIP_SHEETS))


# ── E-Way Bill Consolidator ─────────────────────────────────────
elif ACTIVE == "E-Way Bill Consolidator":
    page_header("📦", "E-Way Bill Consolidator", "Upload multiple Chattisgarh E-Way Bill files and merge them into one formatted Excel sheet.")


    # ─── Page Config ────────────────────────────────────────────────────────────

    # ─── Custom CSS ─────────────────────────────────────────────────────────────
    st.markdown("""
    <style>
      @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

      html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

      /* Dark gradient background */
      .stApp {
        background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
        color: #e2e8f0;
      }

      /* Sidebar */
      section[data-testid="stSidebar"] {
        background: rgba(255,255,255,0.05);
        backdrop-filter: blur(12px);
        border-right: 1px solid rgba(255,255,255,0.1);
      }

      /* Cards */
      .glass-card {
        background: rgba(255,255,255,0.07);
        border: 1px solid rgba(255,255,255,0.12);
        border-radius: 16px;
        padding: 24px;
        margin-bottom: 20px;
        backdrop-filter: blur(10px);
        transition: box-shadow 0.3s ease;
      }
      .glass-card:hover { box-shadow: 0 8px 32px rgba(99,179,237,0.2); }

      /* Hero banner */
      .hero-banner {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 20px;
        padding: 36px 32px;
        margin-bottom: 28px;
        text-align: center;
        box-shadow: 0 20px 60px rgba(102,126,234,0.4);
      }
      .hero-banner h1 {
        font-size: 2.4rem;
        font-weight: 700;
        color: #fff;
        margin: 0 0 8px 0;
        letter-spacing: -0.5px;
      }
      .hero-banner p {
        font-size: 1.05rem;
        color: rgba(255,255,255,0.85);
        margin: 0;
      }

      /* Stat pills */
      .stat-row { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 20px; }
      .stat-pill {
        background: rgba(255,255,255,0.08);
        border: 1px solid rgba(255,255,255,0.15);
        border-radius: 50px;
        padding: 10px 22px;
        font-size: 0.9rem;
        color: #a5d8ff;
        font-weight: 500;
        display: flex;
        align-items: center;
        gap: 8px;
      }

      /* File count badge */
      .badge {
        background: linear-gradient(135deg, #667eea, #764ba2);
        color: #fff;
        border-radius: 20px;
        padding: 4px 12px;
        font-size: 0.8rem;
        font-weight: 600;
      }

      /* Success / warning boxes */
      .success-box {
        background: rgba(72,187,120,0.15);
        border: 1px solid rgba(72,187,120,0.4);
        border-radius: 12px;
        padding: 16px 20px;
        color: #9ae6b4;
        font-weight: 500;
      }
      .warning-box {
        background: rgba(246,173,85,0.15);
        border: 1px solid rgba(246,173,85,0.4);
        border-radius: 12px;
        padding: 16px 20px;
        color: #fbd38d;
        font-weight: 500;
      }
      .error-box {
        background: rgba(252,129,74,0.15);
        border: 1px solid rgba(252,129,74,0.4);
        border-radius: 12px;
        padding: 16px 20px;
        color: #feb2b2;
        font-weight: 500;
      }

      /* Download button */
      .stDownloadButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 14px 36px;
        font-size: 1rem;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.3s ease;
        box-shadow: 0 4px 20px rgba(102,126,234,0.5);
        width: 100%;
      }
      .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(102,126,234,0.7);
      }

      /* File uploader */
      [data-testid="stFileUploader"] {
        background: rgba(255,255,255,0.04);
        border: 2px dashed rgba(102,126,234,0.5);
        border-radius: 16px;
        padding: 12px;
      }

      /* Dataframe */
      [data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }

      /* Progress */
      .stProgress > div > div { background: linear-gradient(90deg, #667eea, #764ba2); }

      /* Divider */
      hr { border-color: rgba(255,255,255,0.1); }

      /* Sidebar headers */
      .sidebar-header {
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 1.5px;
        text-transform: uppercase;
        color: #a0aec0;
        margin: 20px 0 8px 0;
      }
    </style>
    """, unsafe_allow_html=True)

    # ─── Helpers ────────────────────────────────────────────────────────────────

    EXPECTED_COLUMNS = [
        'EWB.No', 'EWB Date', 'Supply Type', 'Sub Supply Type', 'Doc No',
        'Doc Date', 'Doc Type', 'Generated By', 'From GSTIN', 'From Trader Name',
        'From Address', 'From Place', 'From Pincode', 'To GSTIN', 'To Trader Name',
        'To Address', 'To Place', 'To Pincode', 'Status', 'No of Items',
        'Main HSN Code', 'HSN Desc', 'Assessable Value', 'SGST Value',
        'CGST Value', 'IGST Value', 'CESS Value', 'CESS Non.Advol Value',
        'Other Value', 'Total Inv.Value', 'Valid Till Date', 'Mode of GEN.',
        'Cancelled By', 'Cancelled Date', 'IRN', 'Source'
    ]


    def parse_ewb_file(file_content: bytes, filename: str):
        """Parse an E-Way Bill HTML-XLS file and return a DataFrame."""
        try:
            text = file_content.decode('utf-8', errors='ignore')
            soup = BeautifulSoup(text, 'html.parser')
            table = soup.find('table')
            if not table:
                return None, f"No table found in {filename}"

            rows = table.find_all('tr')
            if len(rows) < 2:
                return None, f"No data rows in {filename}"

            # Header row
            headers = [
                cell.get_text(strip=True)
                for cell in rows[0].find_all(['th', 'td'])
            ]

            # Data rows
            data = []
            for row in rows[1:]:
                cells = row.find_all('td')
                values = [cell.get_text(strip=True) for cell in cells]
                # Replace &nbsp; placeholders (already stripped as empty string)
                if len(values) == len(headers):
                    data.append(values)
                elif values:  # partial row – pad
                    values += [''] * (len(headers) - len(values))
                    data.append(values[:len(headers)])

            if not data:
                return None, f"No data rows found in {filename}"

            df = pd.DataFrame(data, columns=headers)
            return df, None

        except Exception as e:
            return None, str(e)


    def build_excel(df: pd.DataFrame, source_breakdown: dict) -> bytes:
        """Build a styled Excel workbook and return bytes."""
        wb = Workbook()

        # ── Sheet 1 : Consolidated Data ──────────────────────────────────────
        ws = wb.active
        ws.title = "Consolidated EWB"

        # Styles
        header_fill = PatternFill("solid", fgColor="1a1a2e")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        alt_fill_1 = PatternFill("solid", fgColor="EEF2FF")   # light lavender
        alt_fill_2 = PatternFill("solid", fgColor="FFFFFF")   # white
        data_font   = Font(name="Calibri", size=9)
        data_align  = Alignment(horizontal="left", vertical="center")

        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Write header
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        # Write data rows
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            fill = alt_fill_1 if row_idx % 2 == 0 else alt_fill_2
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.font = data_font
                cell.alignment = data_align
                cell.border = border

        # Auto-column widths (capped)
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) else 0
            )
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        # ── Sheet 2 : Summary ────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")

        summary_header_fill = PatternFill("solid", fgColor="2D3748")
        summary_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

        # Title
        ws2.merge_cells("A1:D1")
        title_cell = ws2["A1"]
        title_cell.value = "E-Way Bill Consolidation Summary"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="553C9A")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 30

        # Sub-headers
        for col, label in enumerate(["File Name", "Records", "% of Total", "Status"], start=1):
            cell = ws2.cell(row=2, column=col, value=label)
            cell.fill = summary_header_fill
            cell.font = summary_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        total_records = len(df)
        for r_idx, (fname, count) in enumerate(source_breakdown.items(), start=3):
            pct = f"{count / total_records * 100:.1f}%" if total_records else "0%"
            row_fill = PatternFill("solid", fgColor="EBF8FF") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for c_idx, val in enumerate([fname, count, pct, "✅ Processed"], start=1):
                c = ws2.cell(row=r_idx, column=c_idx, value=val)
                c.fill = row_fill
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

        # Totals row
        total_row = len(source_breakdown) + 3
        ws2.merge_cells(f"A{total_row}:A{total_row}")
        for c_idx, val in enumerate(["TOTAL", total_records, "100%", ""], start=1):
            c = ws2.cell(row=total_row, column=c_idx, value=val)
            c.fill = PatternFill("solid", fgColor="553C9A")
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        for col_letter in ["A", "B", "C", "D"]:
            ws2.column_dimensions[col_letter].width = 35 if col_letter == "A" else 18

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


    # ─── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("""
        <div style='text-align:center; padding: 20px 0 10px;'>
          <div style='font-size:3rem;'>📦</div>
          <div style='font-size:1.1rem; font-weight:700; color:#e2e8f0; margin-top:6px;'>EWB Consolidator</div>
          <div style='font-size:0.75rem; color:#a0aec0; margin-top:2px;'>Chattisgarh • E-Way Bills</div>
        </div>
        <hr>
        """, unsafe_allow_html=True)

        st.markdown("<div class='sidebar-header'>⚙️ Options</div>", unsafe_allow_html=True)

        add_source_col = st.checkbox(
            "Add Source File column",
            value=True,
            help="Appends the original filename as the last column"
        )

        remove_duplicates = st.checkbox(
            "Remove duplicate EWB numbers",
            value=False,
            help="Keeps only first occurrence of each EWB.No"
        )

        st.markdown("<div class='sidebar-header'>📋 About</div>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-size:0.82rem; color:#a0aec0; line-height:1.7;'>
        Upload the E-Way Bill <code>.xls</code> files (HTML format exported from the portal).
        The app reads all 36 columns and merges them into a single, styled Excel sheet.
        <br><br>
        <b>Supported format:</b> Portal-generated HTML-XLS files.
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown(
            "<div style='font-size:0.75rem; color:#718096; text-align:center;'>Made with ❤️ for UK APP</div>",
            unsafe_allow_html=True
        )

    # ─── Main Layout ────────────────────────────────────────────────────────────
    st.markdown("""
    <div class='hero-banner'>
      <h1>📦 E-Way Bill Consolidator</h1>
      <p>Upload multiple Chattisgarh E-Way Bill files &amp; merge them into one perfectly formatted Excel sheet</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Upload section ────────────────────────────────────────────────────────
    st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
    st.markdown("### 📁 Upload E-Way Bill Files")
    st.markdown("Select one or more `.xls` files exported from the E-Way Bill portal.")

    uploaded_files = st.file_uploader(
        label="Drop files here or click to browse",
        type=["xls", "xlsx"],
        accept_multiple_files=True,
        help="HTML-format XLS files downloaded from the E-Way Bill portal"
    )
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Process ───────────────────────────────────────────────────────────────
    if uploaded_files:
        st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
        st.markdown(f"### 🔄 Processing <span class='badge'>{len(uploaded_files)} file(s)</span>", unsafe_allow_html=True)

        all_dfs = []
        errors = []
        source_breakdown = {}
        progress = st.progress(0)
        status_placeholder = st.empty()

        for i, uf in enumerate(uploaded_files):
            status_placeholder.markdown(
                f"<div style='color:#a5d8ff; font-size:0.9rem;'>⏳ Parsing: <strong>{uf.name}</strong></div>",
                unsafe_allow_html=True
            )
            content = uf.read()
            df, err = parse_ewb_file(content, uf.name)

            if err:
                errors.append(f"**{uf.name}**: {err}")
            else:
                if add_source_col:
                    df.insert(len(df.columns), "Source File", uf.name)
                all_dfs.append(df)
                source_breakdown[uf.name] = len(df)

            progress.progress((i + 1) / len(uploaded_files))

        status_placeholder.empty()
        st.markdown("</div>", unsafe_allow_html=True)

        # Show errors if any
        if errors:
            st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
            st.markdown("### ⚠️ Warnings")
            for e in errors:
                st.markdown(f"<div class='error-box'>❌ {e}</div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        # ── Stats & Preview ────────────────────────────────────────────────
        if all_dfs:
            consolidated = pd.concat(all_dfs, ignore_index=True)

            if remove_duplicates and 'EWB.No' in consolidated.columns:
                before = len(consolidated)
                consolidated.drop_duplicates(subset=['EWB.No'], keep='first', inplace=True)
                after = len(consolidated)
                if before != after:
                    st.markdown(
                        f"<div class='warning-box'>⚠️ Removed <strong>{before - after}</strong> duplicate EWB numbers. Kept first occurrence.</div>",
                        unsafe_allow_html=True
                    )

            total_records = len(consolidated)
            total_files   = len(all_dfs)
            total_cols    = len(consolidated.columns)

            # Stat pills
            st.markdown(f"""
            <div class='stat-row'>
              <div class='stat-pill'>📄 {total_files} files merged</div>
              <div class='stat-pill'>📊 {total_records:,} total records</div>
              <div class='stat-pill'>🏛️ {total_cols} columns</div>
            </div>
            """, unsafe_allow_html=True)

            # Preview
            st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
            st.markdown("### 👁️ Data Preview (first 50 rows)")
            preview_df = consolidated.head(50)
            st.dataframe(preview_df, use_container_width=True, height=360)
            st.markdown("</div>", unsafe_allow_html=True)

            # Per-file breakdown
            st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
            st.markdown("### 📂 File Breakdown")
            bd_df = pd.DataFrame([
                {
                    "File Name": k,
                    "Records": v,
                    "% of Total": f"{v/total_records*100:.1f}%"
                }
                for k, v in source_breakdown.items()
            ])
            st.dataframe(bd_df, use_container_width=True, hide_index=True)
            st.markdown("</div>", unsafe_allow_html=True)

            # ── Build & Download Excel ─────────────────────────────────────
            st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
            st.markdown("### ⬇️ Download Consolidated Excel")

            with st.spinner("Building Excel workbook…"):
                excel_bytes = build_excel(consolidated, source_breakdown)

            st.markdown(
                f"<div class='success-box'>✅ Ready! <strong>{total_records:,} records</strong> from "
                f"<strong>{total_files} files</strong> consolidated into one sheet.</div>",
                unsafe_allow_html=True
            )
            st.markdown("<br>", unsafe_allow_html=True)

            st.download_button(
                label="⬇️  Download Consolidated E-Way Bills (.xlsx)",
                data=excel_bytes,
                file_name="Consolidated_EWB_Chattisgarh.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.markdown("</div>", unsafe_allow_html=True)

        else:
            st.markdown(
                "<div class='error-box'>❌ No valid data could be extracted from the uploaded files. "
                "Please check that the files are E-Way Bill portal exports.</div>",
                unsafe_allow_html=True
            )

    else:
        # Empty state
        st.markdown("""
        <div class='glass-card' style='text-align:center; padding:48px;'>
          <div style='font-size:4rem; margin-bottom:16px;'>📤</div>
          <div style='font-size:1.2rem; font-weight:600; color:#e2e8f0; margin-bottom:8px;'>
            No files uploaded yet
          </div>
          <div style='font-size:0.9rem; color:#a0aec0;'>
            Use the uploader above to select your E-Way Bill <code>.xls</code> files.<br>
            You can upload multiple files at once — all months, all parts!
          </div>
        </div>
        """, unsafe_allow_html=True)


# ── Excel Consolidator ──────────────────────────────────────────
elif ACTIVE == "Excel Consolidator":
    page_header("📂", "Excel File Consolidator", "Merge multiple Excel / CSV files of any format into one clean, formatted spreadsheet.")

    warnings.filterwarnings("ignore")


    # ── Custom CSS ────────────────────────────────────────────────────────────────
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

        html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

        .main { background: #f8fafc; }

        .hero {
            background: linear-gradient(135deg, #1e3a5f 0%, #2d6a9f 50%, #1a8cff 100%);
            border-radius: 16px;
            padding: 2.5rem 2rem;
            margin-bottom: 1.5rem;
            color: white;
            box-shadow: 0 8px 32px rgba(30,58,95,0.18);
        }
        .hero h1 { font-size: 2.1rem; font-weight: 700; margin: 0 0 0.3rem; letter-spacing: -0.5px; }
        .hero p  { font-size: 1rem; opacity: 0.88; margin: 0; }

        .stat-card {
            background: white;
            border-radius: 12px;
            padding: 1.2rem 1.5rem;
            box-shadow: 0 2px 12px rgba(0,0,0,0.07);
            border-left: 4px solid #2d6a9f;
            margin-bottom: 0.5rem;
        }
        .stat-card .label { font-size: 0.78rem; color: #64748b; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em; }
        .stat-card .value { font-size: 1.6rem; font-weight: 700; color: #1e3a5f; margin-top: 0.1rem; }
        .stat-card .sub   { font-size: 0.82rem; color: #94a3b8; }

        .file-badge {
            display: inline-block;
            padding: 2px 10px;
            border-radius: 20px;
            font-size: 0.75rem;
            font-weight: 600;
            margin: 2px;
        }
        .badge-html { background: #dbeafe; color: #1d4ed8; }
        .badge-xls  { background: #dcfce7; color: #166534; }
        .badge-xlsx { background: #fef9c3; color: #854d0e; }
        .badge-csv  { background: #fce7f3; color: #9d174d; }

        .step-box {
            background: white;
            border-radius: 10px;
            padding: 1rem 1.2rem;
            margin-bottom: 0.8rem;
            box-shadow: 0 1px 6px rgba(0,0,0,0.06);
            border: 1px solid #e2e8f0;
        }
        .step-num {
            display: inline-block;
            width: 26px; height: 26px;
            background: #2d6a9f;
            color: white;
            border-radius: 50%;
            text-align: center;
            line-height: 26px;
            font-size: 0.8rem;
            font-weight: 700;
            margin-right: 8px;
        }

        .success-box {
            background: #f0fdf4;
            border: 1.5px solid #86efac;
            border-radius: 10px;
            padding: 1rem 1.2rem;
            margin: 1rem 0;
        }
        .warning-box {
            background: #fffbeb;
            border: 1.5px solid #fcd34d;
            border-radius: 10px;
            padding: 0.8rem 1.2rem;
            margin: 0.5rem 0;
        }
        .error-box {
            background: #fef2f2;
            border: 1.5px solid #fca5a5;
            border-radius: 10px;
            padding: 0.8rem 1.2rem;
            margin: 0.5rem 0;
        }

        div[data-testid="stFileUploader"] {
            border: 2px dashed #93c5fd;
            border-radius: 12px;
            padding: 0.5rem;
            background: #eff6ff;
        }

        .stButton > button {
            background: linear-gradient(135deg, #1e3a5f, #2d6a9f);
            color: white;
            border: none;
            border-radius: 8px;
            padding: 0.55rem 1.5rem;
            font-weight: 600;
            font-size: 0.95rem;
            transition: all 0.2s;
            box-shadow: 0 2px 8px rgba(30,58,95,0.2);
        }
        .stButton > button:hover { transform: translateY(-1px); box-shadow: 0 4px 14px rgba(30,58,95,0.3); }

        .stDownloadButton > button {
            background: linear-gradient(135deg, #166534, #16a34a) !important;
            color: white !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: 600 !important;
            width: 100%;
        }

        [data-testid="stSidebar"] { background: #1e3a5f; }
        [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
        [data-testid="stSidebar"] .stSelectbox label,
        [data-testid="stSidebar"] .stCheckbox label { color: #cbd5e1 !important; }

        .dataframe thead tr th {
            background: #1e3a5f !important;
            color: white !important;
            font-weight: 600 !important;
        }
    </style>
    """, unsafe_allow_html=True)


    # ── Helpers ───────────────────────────────────────────────────────────────────

    def detect_format(file_bytes: bytes, filename: str) -> str:
        """Detect if file is real XLS (OLE2), HTML-as-XLS, XLSX, or CSV."""
        ext = Path(filename).suffix.lower()
        if ext == ".csv":
            return "csv"
        if ext in (".xlsx", ".xlsm"):
            return "xlsx"
        # Peek at magic bytes
        if file_bytes[:4] == b"\xd0\xcf\x11\xe0":
            return "xls_real"
        # Likely HTML disguised as XLS
        snippet = file_bytes[:200].decode("utf-8", errors="ignore").lower()
        if "<table" in snippet or "<html" in snippet or "<style" in snippet:
            return "xls_html"
        # Try anyway
        return "xls_real"


    def read_html_xls(file_bytes: bytes, source_name: str) -> tuple[pd.DataFrame, str]:
        content = file_bytes.decode("utf-8", errors="ignore")
        soup = BeautifulSoup(content, "html.parser")
        table = soup.find("table")
        if not table:
            raise ValueError("No HTML table found in file.")
        header_row = table.find("tr")
        headers = [th.get_text(strip=True) for th in header_row.find_all(["th", "td"])]
        rows = []
        for tr in table.find_all("tr")[1:]:
            cells = [td.get_text(separator=" ", strip=True) for td in tr.find_all("td")]
            if any(c.strip() for c in cells):
                rows.append(cells)
        if not rows:
            raise ValueError("Table found but no data rows.")
        ncols = max(len(r) for r in rows)
        headers = (headers + [""] * ncols)[:ncols]
        df = pd.DataFrame(rows, columns=headers)
        return df, "HTML-as-XLS"


    def read_real_xls(file_bytes: bytes, source_name: str, sheet_choice: str | None) -> tuple[pd.DataFrame, str]:
        bio = BytesIO(file_bytes)
        sheets = pd.read_excel(bio, engine="xlrd", sheet_name=None, dtype=str)
        if not sheets:
            raise ValueError("No sheets found.")
        if sheet_choice and sheet_choice in sheets:
            df = sheets[sheet_choice]
        else:
            df = list(sheets.values())[0]
        df = df.dropna(how="all").reset_index(drop=True)
        return df, "XLS (Binary)"


    def read_xlsx(file_bytes: bytes, source_name: str, sheet_choice: str | None) -> tuple[pd.DataFrame, str]:
        bio = BytesIO(file_bytes)
        sheets = pd.read_excel(bio, engine="openpyxl", sheet_name=None, dtype=str)
        if not sheets:
            raise ValueError("No sheets found.")
        if sheet_choice and sheet_choice in sheets:
            df = sheets[sheet_choice]
        else:
            df = list(sheets.values())[0]
        df = df.dropna(how="all").reset_index(drop=True)
        return df, "XLSX"


    def read_csv(file_bytes: bytes, source_name: str, encoding: str = "utf-8") -> tuple[pd.DataFrame, str]:
        for enc in [encoding, "latin-1", "cp1252"]:
            try:
                df = pd.read_csv(BytesIO(file_bytes), dtype=str, encoding=enc)
                df = df.dropna(how="all").reset_index(drop=True)
                return df, "CSV"
            except Exception:
                continue
        raise ValueError("Could not decode CSV file.")


    def get_sheet_names(file_bytes: bytes, filename: str) -> list[str]:
        ext = Path(filename).suffix.lower()
        try:
            if ext in (".xlsx", ".xlsm"):
                bio = BytesIO(file_bytes)
                wb = pd.ExcelFile(bio, engine="openpyxl")
                return wb.sheet_names
            elif ext in (".xls",):
                fmt = detect_format(file_bytes, filename)
                if fmt == "xls_real":
                    bio = BytesIO(file_bytes)
                    wb = pd.ExcelFile(bio, engine="xlrd")
                    return wb.sheet_names
        except Exception:
            pass
        return []


    def parse_file(file_bytes: bytes, filename: str, sheet_choice: str | None = None) -> tuple[pd.DataFrame, str]:
        fmt = detect_format(file_bytes, filename)
        if fmt == "csv":
            return read_csv(file_bytes, filename)
        elif fmt == "xlsx":
            return read_xlsx(file_bytes, filename, sheet_choice)
        elif fmt == "xls_html":
            return read_html_xls(file_bytes, filename)
        else:
            return read_real_xls(file_bytes, filename, sheet_choice)


    def build_excel_output(combined: pd.DataFrame, include_summary: bool) -> bytes:
        wb = Workbook()

        H_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        H_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        D_FONT   = Font(name="Arial", size=9)
        ALT_FILL = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
        SRC_FILL = PatternFill("solid", start_color="FFF9C4", end_color="FFF9C4")
        SRC_FONT = Font(name="Arial", size=9, color="7B4F00")
        CENTER   = Alignment(horizontal="center", vertical="center")
        LEFT     = Alignment(horizontal="left",   vertical="center")
        thin     = Side(style="thin", color="CCCCCC")
        BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Sheet 1: All Data ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Consolidated Data"
        cols = list(combined.columns)

        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font   = H_FONT
            cell.fill   = H_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            if col == "Source File":
                cell.fill = PatternFill("solid", start_color="B8860B", end_color="B8860B")

        for ri, row_data in enumerate(combined.itertuples(index=False), 2):
            alt = ri % 2 == 0
            for ci, val in enumerate(row_data, 1):
                v = "" if (val is None or (isinstance(val, float) and np.isnan(val))) else str(val)
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = BORDER
                cell.font   = D_FONT
                col_name = cols[ci - 1]
                if col_name == "Source File":
                    cell.fill      = SRC_FILL
                    cell.font      = SRC_FONT
                    cell.alignment = CENTER
                else:
                    cell.fill      = ALT_FILL if alt else PatternFill()
                    cell.alignment = CENTER if ci <= 3 else LEFT

        # Auto-width (capped)
        for ci, col in enumerate(cols, 1):
            max_len = len(str(col))
            for ri in range(2, min(len(combined) + 2, 200)):
                v = ws.cell(row=ri, column=ci).value or ""
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 45)

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "B2"

        # ── Sheet 2: File Summary ──────────────────────────────────────────────
        if include_summary and "Source File" in combined.columns:
            ws2 = wb.create_sheet("File Summary")
            sum_hdrs = ["Source File", "Rows", "Columns"]
            for ci, h in enumerate(sum_hdrs, 1):
                cell = ws2.cell(row=1, column=ci, value=h)
                cell.font = H_FONT; cell.fill = H_FILL
                cell.alignment = CENTER; cell.border = BORDER

            groups = combined.groupby("Source File", sort=False)
            for ri, (src, grp) in enumerate(groups, 2):
                alt = ri % 2 == 0
                for ci, val in enumerate([src, len(grp), len(combined.columns) - 1], 1):
                    cell = ws2.cell(row=ri, column=ci, value=val)
                    cell.font = D_FONT; cell.alignment = CENTER; cell.border = BORDER
                    if alt: cell.fill = ALT_FILL

            tr = len(groups) + 2
            for ci, val in enumerate(["TOTAL", len(combined), ""], 1):
                cell = ws2.cell(row=tr, column=ci, value=val)
                cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                cell.fill = H_FILL; cell.alignment = CENTER; cell.border = BORDER

            for ci, w in enumerate([40, 12, 12], 1):
                ws2.column_dimensions[get_column_letter(ci)].width = w
            ws2.row_dimensions[1].height = 22

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


    # ── Session state ─────────────────────────────────────────────────────────────
    if "parsed_files" not in st.session_state:
        st.session_state.parsed_files   = {}   # filename -> df
        st.session_state.file_formats   = {}   # filename -> format string
        st.session_state.file_errors    = {}   # filename -> error string
        st.session_state.sheet_choices  = {}   # filename -> chosen sheet
        st.session_state.combined_df    = None


    # ── Sidebar ───────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## ⚙️ Options")
        st.markdown("---")

        output_filename = st.text_input("Output filename", value="Consolidated_Output")
        include_summary = st.checkbox("Add File Summary sheet", value=True)

        st.markdown("### 🧹 Column Handling")
        col_strategy = st.selectbox(
            "Mismatched columns",
            ["Union (keep all columns)", "Intersection (common columns only)"],
        )

        st.markdown("### 🔢 Row Numbering")
        add_row_num = st.checkbox("Add global row number column", value=False)

        st.markdown("---")
        st.markdown("### 📋 Supported Formats")
        for fmt, badge in [
            ("`.xls` — Binary XLS", "badge-xls"),
            ("`.xls` — HTML-as-XLS", "badge-html"),
            ("`.xlsx` / `.xlsm`", "badge-xlsx"),
            ("`.csv`", "badge-csv"),
        ]:
            st.markdown(f'<span class="file-badge {badge}">{fmt}</span>', unsafe_allow_html=True)


    # ── Hero ──────────────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="hero">
        <h1>📊 Excel File Consolidator</h1>
        <p>Merge multiple Excel / CSV files of any format into one clean, formatted spreadsheet — with a <strong>Source File</strong> column tracking every row's origin.</p>
    </div>
    """, unsafe_allow_html=True)


    # ── Step 1: Upload ─────────────────────────────────────────────────────────────
    st.markdown('<div class="step-box"><span class="step-num">1</span> <b>Upload your files</b></div>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Drop files here (XLS, XLSX, CSV — any mix)",
        type=["xls", "xlsx", "xlsm", "csv"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if uploaded:
        st.markdown(f"**{len(uploaded)} file(s) selected**")

        # Sheet selector for multi-sheet files
        sheet_selections = {}
        for uf in uploaded:
            file_bytes = uf.read()
            uf.seek(0)
            sheets = get_sheet_names(file_bytes, uf.name)
            if len(sheets) > 1:
                with st.expander(f"📋 `{uf.name}` has {len(sheets)} sheets — choose one"):
                    chosen = st.selectbox(
                        f"Sheet for {uf.name}",
                        sheets,
                        key=f"sheet_{uf.name}",
                        label_visibility="collapsed",
                    )
                    sheet_selections[uf.name] = chosen

        # ── Step 2: Parse ──────────────────────────────────────────────────────────
        st.markdown('<div class="step-box"><span class="step-num">2</span> <b>Parse & preview files</b></div>', unsafe_allow_html=True)

        if st.button("🔍 Parse All Files"):
            st.session_state.parsed_files  = {}
            st.session_state.file_formats  = {}
            st.session_state.file_errors   = {}
            st.session_state.combined_df   = None

            progress = st.progress(0)
            status   = st.empty()

            for i, uf in enumerate(uploaded):
                status.text(f"Parsing {uf.name}…")
                try:
                    file_bytes = uf.read()
                    sheet_ch   = sheet_selections.get(uf.name)
                    df, fmt    = parse_file(file_bytes, uf.name, sheet_ch)
                    df = df.dropna(how="all").reset_index(drop=True)
                    st.session_state.parsed_files[uf.name]  = df
                    st.session_state.file_formats[uf.name]  = fmt
                except Exception as e:
                    st.session_state.file_errors[uf.name] = str(e)
                progress.progress((i + 1) / len(uploaded))

            status.empty()
            progress.empty()

        # Show parse results
        if st.session_state.parsed_files:
            for fname, df in st.session_state.parsed_files.items():
                fmt = st.session_state.file_formats.get(fname, "")
                badge_cls = {"HTML-as-XLS": "badge-html", "XLS (Binary)": "badge-xls",
                             "XLSX": "badge-xlsx", "CSV": "badge-csv"}.get(fmt, "badge-xls")
                with st.expander(f"✅  `{fname}` — {len(df):,} rows × {len(df.columns)} cols"):
                    st.markdown(f'<span class="file-badge {badge_cls}">{fmt}</span>', unsafe_allow_html=True)
                    st.dataframe(df.head(5), use_container_width=True, hide_index=True)

        for fname, err in st.session_state.file_errors.items():
            st.markdown(f'<div class="error-box">❌ <b>{fname}</b>: {err}</div>', unsafe_allow_html=True)

        # ── Step 3: Consolidate ────────────────────────────────────────────────────
        if st.session_state.parsed_files:
            st.markdown('<div class="step-box"><span class="step-num">3</span> <b>Consolidate</b></div>', unsafe_allow_html=True)

            if st.button("⚡ Consolidate All Files"):
                dfs = []
                for fname, df in st.session_state.parsed_files.items():
                    df = df.copy()
                    df.insert(0, "Source File", fname)
                    dfs.append(df)

                # Column alignment
                if col_strategy.startswith("Union"):
                    combined = pd.concat(dfs, ignore_index=True, sort=False)
                else:
                    common = set(dfs[0].columns)
                    for d in dfs[1:]:
                        common &= set(d.columns)
                    common = sorted(common, key=lambda c: list(dfs[0].columns).index(c) if c in dfs[0].columns else 999)
                    combined = pd.concat([d[list(common)] for d in dfs], ignore_index=True, sort=False)

                if add_row_num:
                    combined.insert(1, "Row No.", range(1, len(combined) + 1))

                st.session_state.combined_df = combined

            if st.session_state.combined_df is not None:
                combined = st.session_state.combined_df

                # Stats
                file_count = combined["Source File"].nunique() if "Source File" in combined.columns else len(st.session_state.parsed_files)
                c1, c2, c3, c4 = st.columns(4)
                for col_obj, label, value, sub in [
                    (c1, "Total Rows",    f"{len(combined):,}",         "across all files"),
                    (c2, "Total Columns", f"{len(combined.columns):,}", "in output sheet"),
                    (c3, "Files Merged",  f"{file_count}",              "source files"),
                    (c4, "Output Size",   f"~{len(combined) * len(combined.columns) // 1000}K", "cells"),
                ]:
                    col_obj.markdown(f"""
                    <div class="stat-card">
                        <div class="label">{label}</div>
                        <div class="value">{value}</div>
                        <div class="sub">{sub}</div>
                    </div>""", unsafe_allow_html=True)

                st.markdown("**Preview — first 50 rows**")
                st.dataframe(combined.head(50), use_container_width=True, hide_index=True)

                # ── Step 4: Download ───────────────────────────────────────────────
                st.markdown('<div class="step-box"><span class="step-num">4</span> <b>Download</b></div>', unsafe_allow_html=True)

                dl_col1, dl_col2 = st.columns(2)

                with dl_col1:
                    with st.spinner("Building Excel file…"):
                        xlsx_bytes = build_excel_output(combined, include_summary)
                    st.download_button(
                        label="⬇️ Download as Excel (.xlsx)",
                        data=xlsx_bytes,
                        file_name=f"{output_filename}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                with dl_col2:
                    csv_bytes = combined.to_csv(index=False).encode("utf-8-sig")
                    st.download_button(
                        label="⬇️ Download as CSV",
                        data=csv_bytes,
                        file_name=f"{output_filename}.csv",
                        mime="text/csv",
                    )

    else:
        # Empty state
        st.markdown("""
        <div style="text-align:center; padding: 3rem 1rem; color: #94a3b8;">
            <div style="font-size: 3.5rem; margin-bottom: 1rem;">📂</div>
            <div style="font-size: 1.1rem; font-weight: 500;">Upload your Excel or CSV files above to get started</div>
            <div style="font-size: 0.9rem; margin-top: 0.5rem;">Supports XLS (binary & HTML), XLSX, XLSM, and CSV</div>
        </div>
        """, unsafe_allow_html=True)

    # ── Footer ─────────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        '<p style="text-align:center; color:#94a3b8; font-size:0.8rem;">Excel Consolidator · Handles XLS (binary & HTML-as-XLS), XLSX, XLSM, CSV · Source column auto-added</p>',
        unsafe_allow_html=True,
    )


# ── E-Invoice Consolidator ──────────────────────────────────────
elif ACTIVE == "E-Invoice Consolidator":
    page_header("🔄", "E-Invoice Consolidator", "Upload multiple Excel files, select sheets & header row, and download a consolidated output.")


    st.markdown("""
        <style>
        .main-title { font-size: 2rem; font-weight: 700; color: #1a3c5e; margin-bottom: 0.2rem; }
        .sub-title { font-size: 1rem; color: #555; margin-bottom: 1.5rem; }
        .step-header { font-size: 1.1rem; font-weight: 600; color: #1a3c5e;
                       background: #e8f0fe; padding: 0.5rem 1rem;
                       border-left: 4px solid #1a73e8; border-radius: 4px; margin-bottom: 0.8rem; }
        .info-box { background: #f0f7ff; border: 1px solid #c2d9f7;
                    border-radius: 6px; padding: 0.7rem 1rem; font-size: 0.88rem; color: #1a3c5e; }
        .success-box { background: #e6f9f0; border: 1px solid #a3dfc4;
                       border-radius: 6px; padding: 0.7rem 1rem; font-size: 0.95rem; color: #1a5c3a; }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="main-title">📊 Excel Consolidator</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Upload multiple Excel files, select sheets & header row, and download a consolidated output.</div>', unsafe_allow_html=True)

    # ── STEP 1: Upload ──────────────────────────────────────────────────────────
    st.markdown('<div class="step-header">Step 1 — Upload Excel Files</div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "Drag & drop or browse to upload .xlsx files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="You can upload as many files as needed."
    )

    if not uploaded_files:
        st.markdown('<div class="info-box">⬆️ Please upload one or more .xlsx files to begin.</div>', unsafe_allow_html=True)
        st.stop()

    st.success(f"✅ {len(uploaded_files)} file(s) uploaded: {', '.join(f.name for f in uploaded_files)}")

    # ── STEP 2: Analyze sheets ──────────────────────────────────────────────────
    st.markdown('<div class="step-header">Step 2 — Select Sheets to Consolidate</div>', unsafe_allow_html=True)

    @st.cache_data(show_spinner="Analyzing sheets…")
    def get_all_sheet_names(file_bytes_list):
        seen = {}
        for name, data in file_bytes_list:
            xl = pd.ExcelFile(io.BytesIO(data))
            for s in xl.sheet_names:
                seen.setdefault(s, []).append(name)
        return seen  # {sheet_name: [file1, file2, …]}

    file_bytes_list = [(f.name, f.read()) for f in uploaded_files]
    # Reset file pointers for later use
    for f in uploaded_files:
        f.seek(0)

    sheet_map = get_all_sheet_names(file_bytes_list)
    all_sheet_names = list(sheet_map.keys())

    st.markdown('<div class="info-box">📋 The following sheets were found across your uploaded files. Select which ones to consolidate.</div>', unsafe_allow_html=True)
    st.write("")

    selected_sheets = st.multiselect(
        "Available Sheets (select one or more)",
        options=all_sheet_names,
        default=all_sheet_names[:1] if all_sheet_names else [],
        help="Only sheets that exist in a file will be pulled from that file."
    )

    if selected_sheets:
        for s in selected_sheets:
            files_with_sheet = sheet_map[s]
            st.caption(f"  • **{s}** — found in {len(files_with_sheet)} file(s): {', '.join(files_with_sheet)}")

    # ── STEP 3: Header row ──────────────────────────────────────────────────────
    st.markdown('<div class="step-header">Step 3 — Select Header Row</div>', unsafe_allow_html=True)
    st.markdown('<div class="info-box">📌 Choose which row contains the column headers. Rows above will be skipped.</div>', unsafe_allow_html=True)
    st.write("")

    header_row = st.selectbox(
        "Header Row Number",
        options=list(range(1, 21)),
        index=3,          # default = Row 4
        format_func=lambda x: f"Row {x}"
    )

    # ── STEP 4: Consolidate ─────────────────────────────────────────────────────
    st.markdown('<div class="step-header">Step 4 — Consolidate & Download</div>', unsafe_allow_html=True)

    if not selected_sheets:
        st.warning("⚠️ Please select at least one sheet above.")
        st.stop()

    def consolidate(file_bytes_list, selected_sheets, header_row_1indexed):
        header_row_0 = header_row_1indexed - 1  # 0-indexed for pandas skiprows
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        summary = {}

        for sheet_name in selected_sheets:
            frames = []
            global_header = None

            for fname, data in file_bytes_list:
                xl = pd.ExcelFile(io.BytesIO(data))
                if sheet_name not in xl.sheet_names:
                    continue

                raw = pd.read_excel(
                    io.BytesIO(data),
                    sheet_name=sheet_name,
                    header=None,
                    dtype=str
                )

                if raw.shape[0] <= header_row_0:
                    continue

                headers = raw.iloc[header_row_0].tolist()
                data_rows = raw.iloc[header_row_0 + 1:].copy()
                data_rows.columns = headers
                data_rows = data_rows.dropna(how="all")

                if global_header is None:
                    global_header = headers

                data_rows.insert(0, "Source", fname)
                frames.append(data_rows)

            if not frames:
                continue

            consolidated_df = pd.concat(frames, ignore_index=True)
            summary[sheet_name] = len(consolidated_df)

            # Write to openpyxl sheet
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars

            # Header styling
            header_fill = PatternFill("solid", start_color="1A3C5E")
            header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            col_names = ["Source"] + (global_header if global_header else [])
            for col_idx, col_name in enumerate(col_names, start=1):
                cell = ws.cell(row=1, column=col_idx, value=str(col_name) if pd.notna(col_name) else "")
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            # Data rows
            data_font = Font(name="Arial", size=10)
            for row_idx, row in enumerate(consolidated_df.itertuples(index=False), start=2):
                for col_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = None if (isinstance(val, float) and pd.isna(val)) else val
                    cell.font = data_font
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill("solid", start_color="F0F4FA")

            # Auto-fit column widths (capped)
            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            ws.freeze_panes = "B2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, summary

    if st.button("🔄 Consolidate & Download", type="primary", use_container_width=True):
        with st.spinner("Consolidating… please wait."):
            try:
                output_bytes, summary = consolidate(file_bytes_list, selected_sheets, header_row)

                st.markdown('<div class="success-box">✅ Consolidation complete! Click the button below to download.</div>', unsafe_allow_html=True)
                st.write("")

                # Summary table
                st.markdown("**Summary of consolidated data:**")
                summary_df = pd.DataFrame(
                    [(s, summary.get(s, 0)) for s in selected_sheets],
                    columns=["Sheet Name", "Total Rows Consolidated"]
                )
                st.dataframe(summary_df, use_container_width=True, hide_index=True)

                st.download_button(
                    label="⬇️ Download Consolidated_Output.xlsx",
                    data=output_bytes,
                    file_name="Consolidated_Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"❌ Error during consolidation: {e}")

    st.write("")
    st.caption("🔒 All processing happens on the server running this app. No data is sent to any third party.")


# ── TDS Challan Extractor ───────────────────────────────────────
elif ACTIVE == "TDS Challan Extractor":
    page_header("💸", "TDS Challan PDF Extractor", "Upload ITNS 281 challan receipts — all data extracted into a single Excel sheet.")


    st.markdown("""
    <style>
        .main-header { font-size: 28px; font-weight: 700; color: #1a1a2e; margin-bottom: 4px; }
        .sub-header { font-size: 14px; color: #6c757d; margin-bottom: 24px; }
        .metric-card { background: #f8f9fa; border-radius: 8px; padding: 16px; text-align: center; border: 1px solid #e9ecef; }
        .metric-value { font-size: 28px; font-weight: 700; color: #1a1a2e; }
        .metric-label { font-size: 12px; color: #6c757d; margin-top: 4px; }
        .stDataFrame { border-radius: 8px; }
        div[data-testid="stFileUploader"] { border: 2px dashed #dee2e6; border-radius: 12px; padding: 8px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="main-header">📄 TDS Challan PDF Extractor</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Upload ITNS 281 challan receipts — all data extracted into a single Excel sheet</div>', unsafe_allow_html=True)


    def extract_value(text, label):
        patterns = [
            rf"{re.escape(label)}\s*[:\-]\s*(.+)",
            rf"{re.escape(label)}\s+(.+)",
        ]
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        return ""


    def clean_amount(val):
        val = val.replace("₹", "").replace(",", "").strip()
        match = re.search(r"[\d]+(?:\.\d+)?", val)
        return float(match.group()) if match else 0.0


    def extract_challan_data(pdf_file):
        with pdfplumber.open(pdf_file) as pdf:
            full_text = ""
            for page in pdf.pages:
                full_text += page.extract_text() + "\n"

        data = {}

        fields = {
            "TAN": "TAN",
            "Name": "Name",
            "Assessment Year": "Assessment Year",
            "Financial Year": "Financial Year",
            "Major Head": "Major Head",
            "Minor Head": "Minor Head",
            "Nature of Payment": "Nature of Payment",
            "CIN": "CIN",
            "Mode of Payment": "Mode of Payment",
            "Bank Name": "Bank Name",
            "Bank Reference Number": "Bank Reference Number",
            "Date of Deposit": "Date of Deposit",
            "BSR code": "BSR Code",
            "Challan No": "Challan No",
            "Tender Date": "Tender Date",
        }

        for label, key in fields.items():
            data[key] = extract_value(full_text, label)

        amount_match = re.search(r"Amount \(in Rs\.\)\s*[:\-]?\s*₹?\s*([\d,]+)", full_text)
        data["Amount (Rs.)"] = clean_amount(amount_match.group(1)) if amount_match else 0.0

        amount_words_match = re.search(r"Amount \(in words\)\s*[:\-]?\s*(.+)", full_text)
        data["Amount (in words)"] = amount_words_match.group(1).strip() if amount_words_match else ""

        breakup_fields = {
            "Tax": r"A\s+Tax\s+₹?\s*([\d,]+)",
            "Surcharge": r"B\s+Surcharge\s+₹?\s*([\d,]+)",
            "Cess": r"C\s+Cess\s+₹?\s*([\d,]+)",
            "Interest": r"D\s+Interest\s+₹?\s*([\d,]+)",
            "Penalty": r"E\s+Penalty\s+₹?\s*([\d,]+)",
            "Fee u/s 234E": r"F\s+Fee under section 234E\s+₹?\s*([\d,]+)",
            "Total": r"Total \(A\+B\+C\+D\+E\+F\)\s+₹?\s*([\d,]+)",
        }

        for key, pattern in breakup_fields.items():
            match = re.search(pattern, full_text)
            data[key] = clean_amount(match.group(1)) if match else 0.0

        itns_match = re.search(r"ITNS No\.\s*[:\-]?\s*(\d+)", full_text)
        data["ITNS No."] = itns_match.group(1).strip() if itns_match else ""

        return data


    def create_excel(records):
        wb = Workbook()
        ws = wb.active
        ws.title = "TDS Challans"

        header_fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        sub_fill = PatternFill("solid", start_color="E8F4FD", end_color="E8F4FD")
        sub_font = Font(bold=True, name="Arial", size=9, color="1a1a2e")
        data_font = Font(name="Arial", size=9)
        alt_fill = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="DEE2E6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells("A1:U1")
        ws["A1"] = "TDS CHALLAN DETAILS — KAPSTON SERVICES LIMITED"
        ws["A1"].font = Font(bold=True, name="Arial", size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        main_headers = [
            "Source File", "S.No", "ITNS No.", "TAN", "Name", "Assessment Year", "Financial Year",
            "Nature of Payment", "CIN", "Mode of Payment", "Bank Name",
            "Bank Ref. No.", "Date of Deposit", "BSR Code", "Challan No", "Tender Date",
            "Tax (Rs.)", "Surcharge (Rs.)", "Cess (Rs.)", "Interest (Rs.)",
            "Penalty (Rs.)", "Fee u/s 234E (Rs.)", "Total Amount (Rs.)"
        ]

        ws.merge_cells("A2:A3")
        ws.merge_cells("B2:B3")

        for col_idx, header in enumerate(main_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.cell(row=3, column=col_idx).font = sub_font
            ws.cell(row=3, column=col_idx).fill = sub_fill
            ws.cell(row=3, column=col_idx).alignment = center
            ws.cell(row=3, column=col_idx).border = border

        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 16

        for i, rec in enumerate(records):
            row = i + 4
            fill = alt_fill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
            values = [
                rec.get("_filename", ""),
                i + 1,
                rec.get("ITNS No.", ""),
                rec.get("TAN", ""),
                rec.get("Name", ""),
                rec.get("Assessment Year", ""),
                rec.get("Financial Year", ""),
                rec.get("Nature of Payment", ""),
                rec.get("CIN", ""),
                rec.get("Mode of Payment", ""),
                rec.get("Bank Name", ""),
                rec.get("Bank Reference Number", ""),
                rec.get("Date of Deposit", ""),
                rec.get("BSR Code", ""),
                rec.get("Challan No", ""),
                rec.get("Tender Date", ""),
                rec.get("Tax", 0),
                rec.get("Surcharge", 0),
                rec.get("Cess", 0),
                rec.get("Interest", 0),
                rec.get("Penalty", 0),
                rec.get("Fee u/s 234E", 0),
                rec.get("Total", 0),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = fill
                cell.border = border
                cell.alignment = center if col_idx == 2 else left
                if col_idx >= 17:
                    cell.number_format = '₹#,##0.00'

            ws.row_dimensions[row].height = 18

        total_row = len(records) + 4
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
        ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
        ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        ws.cell(row=total_row, column=1).alignment = center
        ws.merge_cells(f"A{total_row}:P{total_row}")

        for col_idx in range(17, 24):
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
            cell = ws.cell(row=total_row, column=col_idx, value=formula)
            cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
            cell.number_format = '₹#,##0.00'
            cell.alignment = center
            cell.border = border

        ws.row_dimensions[total_row].height = 20

        col_widths = [24, 5, 8, 14, 28, 14, 12, 18, 26, 14, 14,
                      18, 14, 10, 10, 12, 14, 14, 10, 10, 10, 14, 16]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A4"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


    uploaded_files = st.file_uploader(
        "Upload challan PDF files",
        type=["pdf"],
        accept_multiple_files=True,
        help="Upload one or more ITNS 281 TDS challan PDF files"
    )

    if uploaded_files:
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f'<div class="metric-card"><div class="metric-value">{len(uploaded_files)}</div><div class="metric-label">Files Uploaded</div></div>', unsafe_allow_html=True)

        records = []
        errors = []

        with st.spinner("Extracting data from PDFs..."):
            for f in uploaded_files:
                try:
                    data = extract_challan_data(f)
                    data["_filename"] = f.name
                    records.append(data)
                except Exception as e:
                    errors.append((f.name, str(e)))

        if errors:
            for fname, err in errors:
                st.error(f"❌ {fname}: {err}")

        if records:
            total_amount = sum(r.get("Total", 0) for r in records)
            with col2:
                st.markdown(f'<div class="metric-card"><div class="metric-value">{len(records)}</div><div class="metric-label">Extracted Successfully</div></div>', unsafe_allow_html=True)
            with col3:
                st.markdown(f'<div class="metric-card"><div class="metric-value">₹{total_amount:,.0f}</div><div class="metric-label">Total TDS Amount</div></div>', unsafe_allow_html=True)

            st.markdown("### 📊 Extracted Data Preview")

            preview_cols = [
                "Nature of Payment", "CIN", "Challan No", "Date of Deposit",
                "BSR Code", "Tax", "Surcharge", "Cess", "Interest", "Penalty", "Fee u/s 234E", "Total"
            ]
            df = pd.DataFrame(records)
            df.insert(0, "S.No", range(1, len(df) + 1))
            df["File"] = df["_filename"]

            display_cols = ["S.No", "File"] + [c for c in preview_cols if c in df.columns]
            st.dataframe(df[display_cols], use_container_width=True, hide_index=True)

            st.markdown("### 💾 Export to Excel")
            excel_data = create_excel(records)
            st.download_button(
                label="⬇️ Download Excel File",
                data=excel_data,
                file_name="TDS_Challans.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
            )

    else:
        st.info("👆 Upload one or more TDS challan PDF files to get started.")
        st.markdown("**Supported format:** ITNS 281 Challan Receipts from Income Tax Department")


# ── EPF Challan Consolidator ────────────────────────────────────
elif ACTIVE == "EPF Challan Consolidator":
    page_header("🏭", "EPF Challan Consolidator", "Upload any number of EPF Combined Challan PDFs to merge into a formatted Excel report.")


    # ─── Parsing ───────────────────────────────────────────────────────────────────

    def extract_epf_data(pdf_file):
        try:
            with pdfplumber.open(pdf_file) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        except Exception as e:
            return None, f"Could not read PDF: {e}"

        if "EMPLOYEES' PROVIDENT FUND" not in text:
            return None, "Not an EPF challan — required header not found"

        data = {}

        def find(pattern, flags=0):
            m = re.search(pattern, text, flags)
            return m.group(1).strip() if m else ""

        data["TRRN"]   = find(r"TRRN[:\s]*(\d+)")
        data["ECR Id"] = find(r"ECR\s*Id\s*(\d+)")
        data["LIN"]    = find(r"LIN\s*[:\s]*(\d+)")

        m = re.search(
            r"Establishment Code\s*&\s*([A-Z0-9]+)\s+(.+?)\s+Dues for the wage month\s+(\w+)\s+(\d{4})",
            text
        )
        if m:
            data["Establishment Code"] = m.group(1)
            data["Company Name"]       = m.group(2).strip()
            data["Wage Month"]         = f"{m.group(3)} {m.group(4)}"
        else:
            data["Establishment Code"] = ""
            data["Company Name"]       = ""
            data["Wage Month"]         = ""

        m = re.search(r"Address\s*:\s*(.+?)(?=\nEPF|\nTotal)", text, re.DOTALL)
        data["Address"] = re.sub(r'\s+', ' ', m.group(1)).strip() if m else ""

        def parse_int_pair(pattern):
            m = re.search(pattern, text)
            if m:
                return int(m.group(1).replace(",", "")), int(m.group(2).replace(",", ""))
            return "", ""

        a, b = parse_int_pair(r"Total Subscribers\s*:\s*([\d,]+)\s+([\d,]+)")
        data["Total Subscribers EPF"] = a
        data["Total Subscribers EPS"] = b

        a, b = parse_int_pair(r"Total Wages\s*:\s*([\d,]+)\s+([\d,]+)")
        data["Total Wages EPF"] = a
        data["Total Wages EPS"] = b

        def parse_row(label):
            m = re.search(
                label + r"\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)",
                text
            )
            return [int(g.replace(",", "")) for g in m.groups()] if m else [""] * 6

        for prefix, label in [
            ("Admin",    r"Administration Charges"),
            ("Employer", r"Employer.s Share Of"),
            ("Employee", r"Employee.s Share Of"),
        ]:
            vals = parse_row(label)
            for col, v in zip(["A/C.01","A/C.02","A/C.10","A/C.21","A/C.22","Total"], vals):
                data[f"{prefix} {col}"] = v

        m = re.search(r"Grand Total\s*:\s*(.+?)\s+([\d,]+)\s*$", text, re.MULTILINE)
        if m:
            data["Grand Total (Words)"] = m.group(1).strip()
            data["Grand Total"]         = int(m.group(2).replace(",", ""))
        else:
            data["Grand Total (Words)"] = ""
            data["Grand Total"]         = ""

        m = re.search(r"Total remittance by Employer.*?([\d,]+)\s*$", text, re.MULTILINE)
        data["Total Remittance by Employer"] = int(m.group(1).replace(",", "")) if m else ""

        m = re.search(r"Total amount of uploaded ECR.*?([\d,]+)\s*$", text, re.MULTILINE)
        data["Total ECR Amount"] = int(m.group(1).replace(",", "")) if m else ""

        return data, None


    # ─── Excel Builder ─────────────────────────────────────────────────────────────

    def build_excel(records):
        wb = Workbook()
        ws = wb.active
        ws.title = "EPF Challan Summary"

        thin    = Side(style="thin", color="BFBFBF")
        bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)
        center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_al = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        right_al = Alignment(horizontal="right", vertical="center")

        def mcell(row, col, value=None, bold=False, fg="000000", bg=None, align=None, fmt=None, size=9):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
            if bg: c.fill = PatternFill("solid", start_color=bg)
            c.alignment = align or left_al
            c.border    = bdr
            if fmt: c.number_format = fmt
            return c

        INR_FMT = '#,##0;(#,##0);"-"'
        INT_FMT = "#,##0"

        COLUMNS = [
            # (header, width, key, is_currency, is_int)
            ("S.No",                         5,  "S.No",                       False, True),
            ("Establishment\nCode",         20,  "Establishment Code",          False, False),
            ("Company Name",                28,  "Company Name",                False, False),
            ("Address",                     42,  "Address",                     False, False),
            ("Wage\nMonth",                 13,  "Wage Month",                  False, False),
            ("TRRN",                        18,  "TRRN",                        False, False),
            ("ECR Id",                      14,  "ECR Id",                      False, False),
            ("LIN",                         14,  "LIN",                         False, False),
            ("EPF\nSubscribers",            13,  "Total Subscribers EPF",       False, True),
            ("EPS\nSubscribers",            13,  "Total Subscribers EPS",       False, True),
            ("EPF Total\nWages (Rs.)",      16,  "Total Wages EPF",             True,  False),
            ("EPS Total\nWages (Rs.)",      16,  "Total Wages EPS",             True,  False),
            ("Admin\nA/C.01",               12,  "Admin A/C.01",                True,  False),
            ("Admin\nA/C.02",               12,  "Admin A/C.02",                True,  False),
            ("Admin\nA/C.10",               12,  "Admin A/C.10",                True,  False),
            ("Admin\nA/C.21",               12,  "Admin A/C.21",                True,  False),
            ("Admin\nA/C.22",               12,  "Admin A/C.22",                True,  False),
            ("Admin\nTotal (Rs.)",          14,  "Admin Total",                 True,  False),
            ("Employer\nA/C.01",            12,  "Employer A/C.01",             True,  False),
            ("Employer\nA/C.02",            12,  "Employer A/C.02",             True,  False),
            ("Employer\nA/C.10",            12,  "Employer A/C.10",             True,  False),
            ("Employer\nA/C.21",            12,  "Employer A/C.21",             True,  False),
            ("Employer\nA/C.22",            12,  "Employer A/C.22",             True,  False),
            ("Employer\nTotal (Rs.)",       15,  "Employer Total",              True,  False),
            ("Employee\nA/C.01",            12,  "Employee A/C.01",             True,  False),
            ("Employee\nA/C.02",            12,  "Employee A/C.02",             True,  False),
            ("Employee\nA/C.10",            12,  "Employee A/C.10",             True,  False),
            ("Employee\nA/C.21",            12,  "Employee A/C.21",             True,  False),
            ("Employee\nA/C.22",            12,  "Employee A/C.22",             True,  False),
            ("Employee\nTotal (Rs.)",       15,  "Employee Total",              True,  False),
            ("Grand\nTotal (Rs.)",          16,  "Grand Total",                 True,  False),
            ("Grand Total (In Words)",      45,  "Grand Total (Words)",         False, False),
            ("Total Remittance\nby Employer (Rs.)", 18, "Total Remittance by Employer", True, False),
            ("Total ECR\nAmount (Rs.)",     16,  "Total ECR Amount",            True,  False),
        ]

        SECTIONS = [
            ("Establishment Information",  1,  8,  "1F4E79"),
            ("Subscribers & Wages",        9,  12, "1F4E79"),
            ("Administration Charges",     13, 18, "375623"),
            ("Employer's Share",           19, 24, "843C0C"),
            ("Employee's Share",           25, 30, "7030A0"),
            ("Totals",                     31, 34, "1F4E79"),
        ]

        SEC_COLOR_MAP = {}
        for _, s, e, color in SECTIONS:
            for c in range(s, e+1):
                SEC_COLOR_MAP[c] = color

        # Row 1 — Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        c = ws.cell(row=1, column=1, value="EMPLOYEES' PROVIDENT FUND — CHALLAN CONSOLIDATED REPORT")
        c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Row 2 — Section banners
        ws.row_dimensions[2].height = 16
        for name, s, e, color in SECTIONS:
            ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
            c = ws.cell(row=2, column=s, value=name)
            c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c.fill      = PatternFill("solid", start_color=color)
            c.alignment = center
            c.border    = bdr

        # Row 3 — Column headers
        ws.row_dimensions[3].height = 38
        for col_idx, (hdr, width, _, _, _) in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            bg_color = SEC_COLOR_MAP.get(col_idx, "1F4E79")
            c = ws.cell(row=3, column=col_idx, value=hdr)
            c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c.fill      = PatternFill("solid", start_color=bg_color)
            c.alignment = center
            c.border    = bdr

        # Data rows
        ROW_BG = ["FFFFFF", "EBF3FB"]
        for row_idx, rec in enumerate(records, 1):
            excel_row = row_idx + 3
            ws.row_dimensions[excel_row].height = 16
            bg = ROW_BG[row_idx % 2]
            for col_idx, (_, _, key, is_curr, is_int) in enumerate(COLUMNS, 1):
                val  = row_idx if key == "S.No" else rec.get(key, "")
                fmt  = INR_FMT if is_curr else (INT_FMT if is_int else None)
                al   = right_al if (is_curr or is_int) else left_al
                mcell(excel_row, col_idx, val, fg="000000", bg=bg, align=al, fmt=fmt)

        # Total row
        total_row = len(records) + 4
        ws.row_dimensions[total_row].height = 18
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
        c = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
        c.font      = Font(name="Arial", bold=True, size=10, color="1F4E79")
        c.fill      = PatternFill("solid", start_color="FFF2CC")
        c.alignment = center
        c.border    = bdr

        SUM_COLS = set(range(9, 31)) | {31, 33, 34}
        for col_idx, (_, _, _, is_curr, is_int) in enumerate(COLUMNS, 1):
            if col_idx not in SUM_COLS:
                continue
            col_letter = get_column_letter(col_idx)
            formula    = f"=SUM({col_letter}4:{col_letter}{total_row-1})"
            fmt        = INR_FMT if is_curr else (INT_FMT if is_int else None)
            c = ws.cell(row=total_row, column=col_idx, value=formula)
            c.font      = Font(name="Arial", bold=True, size=9)
            c.fill      = PatternFill("solid", start_color="FFF2CC")
            c.alignment = right_al
            c.border    = bdr
            if fmt: c.number_format = fmt

        ws.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf


    # ─── Streamlit UI ──────────────────────────────────────────────────────────────

    st.title("📋 EPF Challan Consolidator")
    st.markdown(
        "Upload any number of **EPF Combined Challan PDFs** to merge them into a "
        "single formatted Excel report. Non-EPF files are flagged automatically."
    )
    st.markdown("---")

    uploaded = st.file_uploader(
        "Upload EPF Challan PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        help="EPFO Combined Challan (A/C No. 01, 02, 10, 21 & 22)"
    )

    if uploaded:
        valid, results = [], []

        for f in uploaded:
            data, err = extract_epf_data(f)
            if err:
                results.append({"file": f.name, "ok": False, "detail": err})
            else:
                valid.append(data)
                gt     = data.get("Grand Total", "")
                gt_str = f"₹{gt:,}" if isinstance(gt, int) else str(gt)
                detail = (
                    f"Estab: **{data.get('Establishment Code','')}** | "
                    f"Month: **{data.get('Wage Month','')}** | "
                    f"Grand Total: **{gt_str}**"
                )
                results.append({"file": f.name, "ok": True, "detail": detail})

        st.markdown("### 📂 File Processing Results")
        for r in results:
            icon  = "✅" if r["ok"] else "❌"
            label = "Valid EPF Challan" if r["ok"] else "Wrong Format"
            c1, c2, c3 = st.columns([3, 2, 6])
            c1.write(f"`{r['file']}`")
            c2.write(f"{icon} {label}")
            c3.markdown(r["detail"])

        st.markdown("---")

        if valid:
            st.success(f"✅ **{len(valid)} valid challan(s)** ready to consolidate.")

            if st.button("📥 Generate Consolidated Excel", type="primary", use_container_width=True):
                with st.spinner("Building Excel report..."):
                    excel_buf = build_excel(valid)

                st.download_button(
                    label="⬇️ Download EPF_Challan_Consolidated.xlsx",
                    data=excel_buf,
                    file_name="EPF_Challan_Consolidated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                # KPI cards
                st.markdown("### 📊 Summary")
                total_subs  = sum(r.get("Total Subscribers EPF", 0) or 0 for r in valid)
                total_wages = sum(r.get("Total Wages EPF", 0) or 0 for r in valid)
                total_emp   = sum(r.get("Employee Total", 0) or 0 for r in valid)
                total_empr  = sum(r.get("Employer Total", 0) or 0 for r in valid)
                total_grand = sum(r.get("Grand Total", 0) or 0 for r in valid)

                k1, k2, k3, k4, k5 = st.columns(5)
                k1.metric("Establishments",   len(valid))
                k2.metric("Total EPF Members", f"{total_subs:,}")
                k3.metric("Total Wages",       f"Rs.{total_wages:,.0f}")
                k4.metric("Employee Share",    f"Rs.{total_emp:,.0f}")
                k5.metric("Grand Total",       f"Rs.{total_grand:,.0f}")

                st.markdown("### 🏢 Per-Establishment Details")
                for i, r in enumerate(valid, 1):
                    gt = r.get("Grand Total", 0) or 0
                    with st.expander(
                        f"{i}. {r.get('Establishment Code','')} — "
                        f"{r.get('Company Name','')} | {r.get('Wage Month','')} | "
                        f"Rs.{gt:,}"
                    ):
                        col_a, col_b, col_c = st.columns(3)
                        col_a.markdown(f"**Address:** {r.get('Address','')}")
                        col_a.markdown(f"**TRRN:** `{r.get('TRRN','')}`")
                        col_a.markdown(f"**LIN:** `{r.get('LIN','')}`")
                        col_b.metric("EPF Subscribers", r.get("Total Subscribers EPF",""))
                        col_b.metric("EPS Subscribers", r.get("Total Subscribers EPS",""))
                        col_b.metric("Total Wages", f"Rs.{r.get('Total Wages EPF',0):,}")
                        col_c.metric("Admin Charges",  f"Rs.{r.get('Admin Total',0):,}")
                        col_c.metric("Employer Share", f"Rs.{r.get('Employer Total',0):,}")
                        col_c.metric("Employee Share", f"Rs.{r.get('Employee Total',0):,}")
        else:
            st.error("No valid EPF challans found. Please upload correct EPFO Combined Challan PDFs.")

    else:
        st.info("👆 Upload one or more EPF Challan PDFs above to get started.")
        with st.expander("ℹ️ How it works"):
            st.markdown("""
            1. **Upload** — Drop any number of EPFO Combined Challan PDFs
            2. **Parse** — Each PDF is automatically validated and parsed
            3. **Export** — Click Generate to download a consolidated Excel with:
               - All establishment details (code, address, TRRN, LIN)
               - Subscriber and wage counts
               - Admin, Employer & Employee contribution breakdowns (A/C.01–22)
               - Grand totals per challan and overall grand total row
            4. **Wrong Format** — Any non-EPF PDFs are flagged and skipped
            """)


# ── ESI Challan Extractor ───────────────────────────────────────
elif ACTIVE == "ESI Challan Extractor":
    page_header("🏥", "ESIC Challan Extractor", "Upload one or more ESIC Challan PDFs to extract and export data to Excel.")


    st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

        html, body, [class*="css"] {
            font-family: 'IBM Plex Sans', sans-serif;
        }
        .main { background-color: #f5f5f0; }
        .stApp { background-color: #f5f5f0; }

        h1 {
            font-family: 'IBM Plex Mono', monospace !important;
            font-size: 1.6rem !important;
            color: #1a1a2e !important;
            border-bottom: 3px solid #c0392b;
            padding-bottom: 0.4rem;
        }
        .stButton > button {
            background-color: #c0392b;
            color: white;
            border: none;
            border-radius: 4px;
            font-family: 'IBM Plex Mono', monospace;
            font-weight: 600;
            padding: 0.5rem 1.5rem;
            width: 100%;
        }
        .stButton > button:hover {
            background-color: #922b21;
        }
        .upload-box {
            background: white;
            border: 2px dashed #c0392b;
            border-radius: 8px;
            padding: 1rem;
            margin-bottom: 1rem;
        }
        .tag {
            display: inline-block;
            background: #1a1a2e;
            color: white;
            font-family: 'IBM Plex Mono', monospace;
            font-size: 0.75rem;
            padding: 2px 8px;
            border-radius: 3px;
            margin-right: 4px;
        }
        </style>
    """, unsafe_allow_html=True)

    st.title("ESIC Challan Extractor")
    st.markdown("Upload one or more ESIC Challan PDFs to extract and export data to Excel.")

    FIELDS = {
        "Employer's Code No": "Employer Code No",
        "Employer's Name": "Employer Name",
        "Challan Period": "Challan Period",
        "Challan Number": "Challan Number",
        "Challan Created Date": "Challan Created Date",
        "Challan Submitted Date": "Challan Submitted Date",
        "Amount Paid": "Amount Paid",
        "Transaction Number": "Transaction Number",
        "Transaction status": "Transaction Status",
    }

    def extract_from_pdf(file_bytes):
        data = {}
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)

        for field, col in FIELDS.items():
            pattern = re.escape(field) + r"[\s:]*([^\n]+)"
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                data[col] = match.group(1).strip().rstrip("*").strip()
            else:
                data[col] = ""
        return data

    def create_excel(records):
        wb = Workbook()
        ws = wb.active
        ws.title = "ESIC Challans"

        headers = ["Source File"] + list(FIELDS.values())

        # Styles
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", start_color="1A1A2E")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="Arial", size=10)
        alt_fill = PatternFill("solid", start_color="F2F2EF")
        center_align = Alignment(horizontal="center", vertical="center")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        ws.row_dimensions[1].height = 30

        # Data rows
        for row_idx, record in enumerate(records, 2):
            fill = PatternFill("solid", start_color="FFFFFF") if row_idx % 2 == 0 else alt_fill
            for col_idx, header in enumerate(headers, 1):
                val = record.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = fill
                cell.alignment = center_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 20

        # Column widths
        col_widths = [30, 22, 28, 14, 22, 22, 22, 14, 22, 26]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Total row
        total_row = len(records) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="C0392B")
        ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF")
        ws.cell(row=total_row, column=1).alignment = center_align

        amt_col = headers.index("Amount Paid") + 1
        total_formula = f"=SUM({get_column_letter(amt_col)}2:{get_column_letter(amt_col)}{total_row-1})"
        total_cell = ws.cell(row=total_row, column=amt_col, value=total_formula)
        total_cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        total_cell.fill = PatternFill("solid", start_color="C0392B")
        total_cell.alignment = center_align
        total_cell.border = border

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # Upload
    uploaded_files = st.file_uploader(
        "Upload ESIC Challan PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        help="You can upload multiple PDFs at once"
    )

    if uploaded_files:
        st.markdown(f"**{len(uploaded_files)} file(s) uploaded**")

        records = []
        errors = []

        for f in uploaded_files:
            try:
                record = extract_from_pdf(f.read())
                record["_filename"] = f.name
                records.append(record)
            except Exception as e:
                errors.append(f"{f.name}: {e}")

        if errors:
            for err in errors:
                st.error(f"⚠️ {err}")

        if records:
            display_records = [{"Source File": r["_filename"], **{k: v for k, v in r.items() if k != "_filename"}} for r in records]
            df = pd.DataFrame(display_records)

            st.markdown("### Preview")
            st.dataframe(df, use_container_width=True)

            st.markdown(f"**{len(records)} record(s) extracted** | Total Amount: ₹{sum(float(r.get('Amount Paid', 0) or 0) for r in display_records):,.2f}")

            if st.button("⬇ Download Excel"):
                excel_file = create_excel(display_records)
                st.download_button(
                    label="📥 Click to Save Excel File",
                    data=excel_file,
                    file_name="ESIC_Challans.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("Upload PDFs above to get started.")


# ── Ind AS 116 – Lease ──────────────────────────────────────────
elif ACTIVE == "Ind AS 116 – Lease":
    page_header("📐", "Ind AS 116 – Lease Liability Working", "Amortisation schedule · ROU asset · Journal entries · Excel with live formulas")
    """
    Ind AS 116 – Lease Liability Working  (Streamlit)
    Run:   streamlit run IndAS116_Streamlit.py
    Needs: pip install streamlit openpyxl python-dateutil pandas
    """


    st.markdown("""<style>
    [data-testid="stAppViewContainer"]{background:#F7F7F5}
    .block-container{padding:1.4rem 2rem 3rem;max-width:1440px}
    .ttl{background:#1A1A1A;border-radius:10px;padding:13px 22px;margin-bottom:1.1rem}
    .ttl h1{color:#fff;font-size:19px;font-weight:700;margin:0}
    .ttl p{color:#9CA3AF;font-size:12px;margin:2px 0 0}
    .slbl{font-size:10.5px;font-weight:700;color:#6B7280;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.5rem}
    .callout{background:#EFF6FF;border-left:3px solid #3B82F6;border-radius:0 6px 6px 0;padding:9px 13px;font-size:12px;color:#1E40AF;margin:4px 0 10px}
    .camber{background:#FFFBEB;border-left:3px solid #F59E0B;border-radius:0 6px 6px 0;padding:9px 13px;font-size:12px;color:#92400E;margin:4px 0 10px}
    label{font-size:12px !important;font-weight:600 !important;color:#374151 !important}
    div[data-testid="stButton"]>button{background:#1A1A1A !important;color:#fff !important;border:none !important;border-radius:7px !important;font-weight:700 !important;font-size:13px !important;padding:9px 0 !important;width:100%}
    div[data-testid="stButton"]>button:hover{background:#374151 !important}
    div[data-testid="stDownloadButton"]>button{background:#1D9E75 !important;color:#fff !important;border:none !important;border-radius:7px !important;font-weight:700 !important;font-size:13px !important;padding:9px 0 !important;width:100%}
    div[data-testid="stDownloadButton"]>button:hover{background:#178a64 !important}
    .stTabs [data-baseweb="tab-list"]{gap:2px;background:#F3F4F6;border-radius:8px 8px 0 0;padding:4px 4px 0}
    .stTabs [data-baseweb="tab"]{font-size:12.5px;font-weight:500;border-radius:6px 6px 0 0;padding:7px 18px;color:#6B7280}
    .stTabs [aria-selected="true"]{background:#fff !important;color:#1A1A1A !important;border-bottom:2px solid #1A1A1A !important}
    div[data-testid="stDataFrame"]{border-radius:8px;overflow:hidden}
    hr{border:none;border-top:1px solid #E5E7EB;margin:.9rem 0}
    </style>""", unsafe_allow_html=True)

    # ── helpers ────────────────────────────────────────────────────────────────────
    def add_months(dt, n): return dt + relativedelta(months=n)
    def inr(n):            return f"₹ {n:,.2f}"

    def parse_date(s):
        for fmt in ("%d-%m-%Y","%d/%m/%Y","%Y-%m-%d","%d-%b-%Y","%d %b %Y"):
            try: return datetime.strptime(s.strip(), fmt).date()
            except ValueError: pass
        raise ValueError(f"Cannot parse '{s}'. Use DD-MM-YYYY.")

    FREQ_LABEL = {1:"Monthly",3:"Quarterly",6:"Half-Yearly",12:"Yearly"}

    # ── calculation engine ─────────────────────────────────────────────────────────
    def compute_schedule(from_date, to_date, lease_payment, annual_rate,
                         months_override, esc_rate, esc_every_n, pay_freq, advance):
        """
        lease_payment : cash amount paid each time (e.g. 50000 per 6-month block)
        pay_freq      : frequency in months (1=monthly, 3=quarterly, 6=half-yearly, 12=yearly)
        advance       : True = paid at START of period, False = paid at END of period

        PV discount:
          advance  → t = k * pay_freq  (first payment at t=0, DF=1)
          arrears  → t = (k+1)*pay_freq (first payment at t=pay_freq)

        Non-payment rows: cash=0, DF=blank, PV=0; interest still accrues on liability.
        Payment rows: interest on (opening – cash) if advance, else interest on opening.
        """
        mr = annual_rate / 100 / 12
        if months_override and int(months_override) > 0:
            nm = int(months_override)
        else:
            rd = relativedelta(to_date, from_date)
            nm = rd.years * 12 + rd.months
            if nm <= 0: raise ValueError("End date must be after start date.")

        n_esc      = max(1, int(esc_every_n))
        freq       = max(1, int(pay_freq))
        num_blocks = math.ceil(nm / freq)

        def block_payment(k):
            # escalation based on the month index when this payment falls
            pay_month = k * freq if advance else min((k+1)*freq - 1, nm-1)
            esc_periods = pay_month // n_esc
            return round(lease_payment * math.pow(1 + esc_rate/100, esc_periods), 2)

        cash_out = [0.0] * nm
        df_map   = {}        # month_index → discount_factor (only on payment months)
        pv       = 0.0

        for k in range(num_blocks):
            pmt     = block_payment(k)
            pay_idx = k * freq if advance else min((k+1)*freq - 1, nm-1)
            t       = pay_idx  if advance else pay_idx + 1   # time in months for discounting
            if pay_idx < nm:
                cash_out[pay_idx] = pmt
                df                = 1 / math.pow(1 + mr, t)
                df_map[pay_idx]   = df
                pv               += pmt * df

        rou_depr = pv / nm
        rows, open_liab = [], pv

        for i in range(nm):
            cash  = cash_out[i]
            df    = df_map.get(i, None)          # None → blank on non-payment rows
            pv_col = round(cash * df, 2) if (df is not None and cash > 0) else 0.0

            if advance and cash > 0:
                # advance payment: deduct cash first, then interest on remainder
                after    = open_liab - cash
                interest = after * mr
                cl       = after + interest
            else:
                # arrears or non-payment month: interest on full opening
                interest = open_liab * mr
                cl       = open_liab + interest - cash

            cl  = 0.0 if abs(cl)  < 0.005 else cl
            rou = pv - rou_depr * (i + 1)
            rou = 0.0 if abs(rou) < 0.005 else rou

            rows.append({
                "sno":    i + 1,
                "date":   add_months(from_date, i + 1),
                "mp":     cash_out[i],
                "df":     df,
                "pv_col": pv_col,
                "cash":   cash,
                "open":   round(open_liab, 2),
                "int":    round(interest, 2),
                "close":  round(cl, 2),
                "depr":   round(rou_depr, 2),
                "rou":    round(rou, 2),
            })
            open_liab = cl

        return rows, pv, mr, nm, rou_depr


    # ── Excel builder ──────────────────────────────────────────────────────────────
    def build_excel(schedule, info):
        """
        ARREARS → 11 cols A–K:
          A:S.No  B:Month  C:Lease Payment  D:Discount Factor  E:Present Value
          F:Opening Liability  G:Interest Expense  H:Lease Payment(cash)
          I:Closing Liability  J:Depreciation  K:ROU Asset

        ADVANCE → 12 cols A–L (extra col D = Advance Cash Paid):
          A:S.No  B:Month  C:Lease Payment  D:Advance Cash Paid  E:Discount Factor
          F:Present Value  G:Opening Liability  H:Interest Expense  I:Lease Payment(cash)
          J:Closing Liability  K:Depreciation  L:ROU Asset

        Key formula rules (no Python operators in Excel strings):
          - Escalation : =$B$10*(1+$B$13/100)^INT((A{r}-1)/$B$14)
          - Discount t  : hardcoded integer exponent derived from df value
          - C non-pay   : hardcoded 0 (arrears, non-payment row)
          - D/E non-pay : empty string (blank cell)
        """
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = info["name"][:31]
        ws.sheet_view.showGridLines = False

        adv   = info["advance"]
        NM    = info["num_months"]
        freq  = info["pay_freq"]
        ar    = info["annual_rate"]
        mr    = ar / 100 / 12
        NF    = "#,##0.00"
        NCOLS = 12 if adv else 11

        thin = Side(border_style="thin", color="D1D5DB")
        def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)

        def W(row, col, val, bold=False, align="left", fmt=None,
              bg=None, fg="1A1A1A", size=10, wrap=False):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", bold=bold, size=size, color=fg)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            if bg:  c.fill          = PatternFill("solid", fgColor=bg)
            if fmt: c.number_format = fmt
            return c

        def MW(r, c1, c2, val, **kw):
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            W(r, c1, val, **kw)

        def t_from_df(df_val):
            """Recover integer t from df = 1/(1+mr)^t"""
            if df_val is None or df_val <= 0:
                return None
            if abs(df_val - 1.0) < 1e-9:
                return 0
            return int(round(-math.log(df_val) / math.log(1 + mr)))

        # ── Rows 1–2: Title & lease name ─────────────────────────────────────────
        ws.row_dimensions[1].height = 30
        MW(1, 1, NCOLS, "IND AS 116 – LEASE LIABILITY WORKING",
           bold=True, align="center", size=14, bg="1A1A1A", fg="FFFFFF")
        ws.row_dimensions[2].height = 22
        MW(2, 1, NCOLS, info["name"],
           bold=True, align="center", size=12, bg="1D9E75", fg="FFFFFF")
        ws.row_dimensions[3].height = 6

        # ── Info block rows 4–14 ─────────────────────────────────────────────────
        freq_label = FREQ_LABEL.get(freq, f"Every {freq} months")
        adv_label  = f"{'Advance' if adv else 'Arrears'} – {freq_label}"
        meta = [
            (4,  "Lease Term (Years)",                    round(NM/12, 4),         "0.00",  "0000FF"),
            (5,  "Beginning From",                        info["from_date"].strftime("%d-%b-%Y"), None, "0000FF"),
            (6,  "Ending To",                             info["to_date"].strftime("%d-%b-%Y"),   None, "0000FF"),
            (7,  "No. of Months",                         NM,                      "0",     "0000FF"),
            (8,  "Annual Discount Rate",                  ar / 100,                "0.00%", "0000FF"),
            (9,  "Monthly Discount Rate",                 "=B8/12",                "0.00%", "1A1A1A"),
            (10, "Lease Payment (Rs.)",                   info["lease_payment"],   NF,      "0000FF"),
            (11, "Payment Frequency",                     freq_label,              None,    "0000FF"),
            (12, "Payment Mode",                          adv_label,               None,    "0000FF"),
            (13, "Escalation Rate (%)",                   info["esc_rate"],         "0.00",  "0000FF"),
            (14, "Escalation Frequency (every N months)", info["esc_every_n"],      "0",     "0000FF"),
        ]
        for (row_n, lbl, val, fmt, fg) in meta:
            ws.row_dimensions[row_n].height = 15
            W(row_n, 1, lbl, bold=True, fg="374151", size=9)
            W(row_n, 2, val, fmt=fmt,   fg=fg,       size=9)

        PV_ROW = 15
        ws.row_dimensions[PV_ROW].height = 16
        W(PV_ROW, 1, "PV of Lease Payments – Initial Lease Liability",
          bold=True, fg="374151", size=9)
        W(PV_ROW, 2, round(info["pv"], 2), fmt=NF, bold=True, size=10, fg="1D9E75")
        ws.row_dimensions[16].height = 8
        ws.row_dimensions[17].height = 8

        # ── Header row 18, data starts row 19 ────────────────────────────────────
        HDR_ROW    = 18
        DATA_START = 19
        PV_REF     = f"$B${PV_ROW}"    # =$B$15
        NM_REF     = "$B$7"            # No. of months

        ws.row_dimensions[HDR_ROW].height = 44
        if not adv:
            HDR_LABELS = [
                "S.No", "Month",
                "Lease\nPayment\n(Rs.)", "Discount\nFactor", "Present\nValue\n(Rs.)",
                "Opening\nLiability\n(Rs.)", "Interest\nExpense\n(Rs.)",
                "Lease\nPayment\n(Rs.)", "Closing\nLiability\n(Rs.)",
                "Depreciation\n(Rs.)", "ROU Asset\n(Rs.)",
            ]
            COL_WIDTHS = [6, 13, 15, 13, 15, 18, 15, 15, 18, 15, 15]
        else:
            HDR_LABELS = [
                "S.No", "Month",
                "Lease\nPayment\n(Rs.)", "Advance\nCash Paid\n(Rs.)",
                "Discount\nFactor", "Present\nValue\n(Rs.)",
                "Opening\nLiability\n(Rs.)", "Interest\nExpense\n(Rs.)",
                "Lease\nPayment\n(Rs.)", "Closing\nLiability\n(Rs.)",
                "Depreciation\n(Rs.)", "ROU Asset\n(Rs.)",
            ]
            COL_WIDTHS = [6, 13, 15, 16, 13, 15, 18, 15, 15, 18, 15, 15]

        for ci, hdr in enumerate(HDR_LABELS, 1):
            c = W(HDR_ROW, ci, hdr, bold=True, align="center",
                   bg="1A1A1A", fg="FFFFFF", size=9, wrap=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr()
        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # ── Data rows ─────────────────────────────────────────────────────────────
        for ri, row in enumerate(schedule):
            r      = DATA_START + ri
            bg     = "F3F4F6" if ri % 2 else None
            is_pay = row["cash"] > 0
            df_val = row["df"]           # None on non-payment rows
            t      = t_from_df(df_val)  # integer discount exponent (None if non-payment)
            ws.row_dimensions[r].height = 15

            def wr(col, val, fmt=None, align="right", _r=r, _bg=bg):
                c2 = W(_r, col, val, fmt=fmt, align=align, bg=_bg, size=9)
                c2.border = bdr()
                return c2

            # ── ESCALATION formula helper (valid Excel, no Python operators) ──────
            # =$B$10*(1+$B$13/100)^INT((A{r}-1)/$B$14)
            # Uses cell A{r} which contains the S.No (row number 1-based)
            esc_formula = f"=$B$10*(1+$B$13/100)^INT((A{r}-1)/$B$14)"

            if not adv:
                # ══════════════════════════════════════════════════════════════════
                # ARREARS – 11 columns A to K
                # ══════════════════════════════════════════════════════════════════
                # A: S.No (hardcoded)
                wr(1, ri + 1, fmt="0", align="center")

                # B: Month (hardcoded date string)
                wr(2, row["date"].strftime("%d-%b-%Y"), align="center")

                # C: Lease Payment
                #    payment row  → escalation formula (driven by B10, B13, B14)
                #    non-payment  → hardcoded 0
                if is_pay:
                    wr(3, esc_formula, fmt=NF)
                else:
                    wr(3, 0, fmt=NF)

                # D: Discount Factor
                #    payment row  → =1/(1+$B$9)^t  where t is integer
                #    non-payment  → blank
                if is_pay and t is not None:
                    wr(4, f"=1/(1+$B$9)^{t}", fmt="0.00000000")
                else:
                    wr(4, None, fmt="0.00000000")

                # E: Present Value = C * D
                #    payment row  → =C{r}*D{r}
                #    non-payment  → blank
                if is_pay:
                    wr(5, f"=C{r}*D{r}", fmt=NF)
                else:
                    wr(5, None, fmt=NF)

                # F: Opening Liability
                #    row 1 → =$B$15 (PV)
                #    rest  → =I{r-1} (previous closing)
                wr(6, f"={PV_REF}" if ri == 0 else f"=I{r-1}", fmt=NF)

                # G: Interest Expense = F * $B$9  (interest on full opening, arrears)
                wr(7, f"=F{r}*$B$9", fmt=NF)

                # H: Lease Payment (cash out) = C  (mirrors C column)
                wr(8, f"=C{r}", fmt=NF)

                # I: Closing Liability = F + G – H
                wr(9, f"=F{r}+G{r}-H{r}", fmt=NF)

                # J: Depreciation = PV / No. of months
                wr(10, f"={PV_REF}/{NM_REF}", fmt=NF)

                # K: ROU Asset = PV – J * S.No
                wr(11, f"={PV_REF}-J{r}*A{r}", fmt=NF)

            else:
                # ══════════════════════════════════════════════════════════════════
                # ADVANCE – 12 columns A to L
                # ══════════════════════════════════════════════════════════════════
                # A: S.No
                wr(1, ri + 1, fmt="0", align="center")

                # B: Month
                wr(2, row["date"].strftime("%d-%b-%Y"), align="center")

                # C: Monthly lease payment equivalent (escalation formula every row)
                wr(3, esc_formula, fmt=NF)

                # D: Advance Cash Paid
                #    payment row  → =SUM(C{r}:C{block_end})  (sum of freq months)
                #    non-payment  → hardcoded 0
                if is_pay:
                    block_end_r = min(r + freq - 1, DATA_START + NM - 1)
                    wr(4, f"=SUM(C{r}:C{block_end_r})", fmt=NF)
                else:
                    wr(4, 0, fmt=NF)

                # E: Discount Factor
                #    payment row  → =1/(1+$B$9)^t
                #    non-payment  → blank
                if is_pay and t is not None:
                    wr(5, f"=1/(1+$B$9)^{t}", fmt="0.00000000")
                else:
                    wr(5, None, fmt="0.00000000")

                # F: Present Value = C * E on payment rows, blank otherwise
                if is_pay:
                    wr(6, f"=C{r}*E{r}", fmt=NF)
                else:
                    wr(6, None, fmt=NF)

                # G: Opening Liability
                wr(7, f"={PV_REF}" if ri == 0 else f"=J{r-1}", fmt=NF)

                # H: Interest Expense = (G – D) * $B$9
                #    (advance: interest on opening after deducting cash paid)
                wr(8, f"=(G{r}-D{r})*$B$9", fmt=NF)

                # I: Lease Payment (cash out) = D  (mirrors D column)
                wr(9, f"=D{r}", fmt=NF)

                # J: Closing Liability = G – D + H
                wr(10, f"=G{r}-D{r}+H{r}", fmt=NF)

                # K: Depreciation = PV / No. of months
                wr(11, f"={PV_REF}/{NM_REF}", fmt=NF)

                # L: ROU Asset = PV – K * S.No
                wr(12, f"={PV_REF}-K{r}*A{r}", fmt=NF)

        # ── Totals row ────────────────────────────────────────────────────────────
        TOT = DATA_START + NM
        ws.row_dimensions[TOT].height = 17
        # Col A: label (not merged → safe)
        c = W(TOT, 1, "TOTALS", bold=True, align="right", bg="374151", fg="FFFFFF", size=9)
        c.border = bdr()
        # Col B: style only (no value — avoids MergedCell read-only error)
        ws.cell(TOT, 2).fill   = PatternFill("solid", fgColor="374151")
        ws.cell(TOT, 2).border = bdr()
        # Cols 3+: SUM formulas for relevant cols, dash for others
        if not adv:
            sum_cols   = {3, 5, 7, 8}       # C=pmt, E=PV, G=interest, H=cash
            total_cols = 11
        else:
            sum_cols   = {3, 4, 6, 8, 9}    # C=pmt, D=adv cash, F=PV, H=interest, I=cash
            total_cols = 12
        for ci in range(3, total_cols + 1):
            col = get_column_letter(ci)
            if ci in sum_cols:
                c = W(TOT, ci, f"=SUM({col}{DATA_START}:{col}{TOT-1})",
                      fmt=NF, bold=True, bg="374151", fg="FFFFFF", size=9, align="right")
            else:
                c = W(TOT, ci, "—", bold=True, bg="374151", fg="9CA3AF", size=9, align="center")
            c.border = bdr()

        # ── Year-1 journal summary ────────────────────────────────────────────────
        SUM_R = TOT + 3
        ws.row_dimensions[SUM_R].height = 18
        MW(SUM_R, 1, NCOLS, "YEAR 1 – JOURNAL ENTRY SUMMARY",
           bold=True, align="center", bg="374151", fg="FFFFFF", size=10)
        yr1_end  = DATA_START + min(12, NM) - 1
        depr_col = "J" if not adv else "K"
        int_col  = "G" if not adv else "H"
        cash_col = "H" if not adv else "D"
        jrows = [
            ("Depreciation A/c",                "Dr", f"=SUM({depr_col}{DATA_START}:{depr_col}{yr1_end})", ""),
            ("   To ROU Asset A/c",             "Cr", "", f"=SUM({depr_col}{DATA_START}:{depr_col}{yr1_end})"),
            ("", "", "", ""),
            ("Interest on Lease Liability A/c", "Dr", f"=SUM({int_col}{DATA_START}:{int_col}{yr1_end})", ""),
            ("   To Lease Liability A/c",       "Cr", "", f"=SUM({int_col}{DATA_START}:{int_col}{yr1_end})"),
            ("", "", "", ""),
            ("Lease Liability A/c",             "Dr", f"=SUM({cash_col}{DATA_START}:{cash_col}{yr1_end})", ""),
            ("   To Bank / Cash A/c",           "Cr", "", f"=SUM({cash_col}{DATA_START}:{cash_col}{yr1_end})"),
        ]
        for ji, (part, dc2, dr, cr) in enumerate(jrows):
            row_n = SUM_R + 2 + ji
            ws.row_dimensions[row_n].height = 14
            W(row_n, 1, part, size=9, fg="111827")
            W(row_n, 2, dc2,  size=9, fg="6B7280", bold=True, align="center")
            if dr: W(row_n, 3, dr, fmt=NF, align="right", size=9)
            if cr: W(row_n, 4, cr, fmt=NF, align="right", size=9)

        ws.freeze_panes = f"A{DATA_START}"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


    # ══════════════════════════════════════════════════════════════════════════════
    # UI
    # ══════════════════════════════════════════════════════════════════════════════
    st.markdown("""<div class="ttl">
      <h1>📋 Ind AS 116 – Lease Liability Working</h1>
      <p>Amortisation schedule · ROU asset · Journal entries · Excel with live formulas</p>
    </div>""", unsafe_allow_html=True)

    # ── Section 1: Lease Details ───────────────────────────────────────────────────
    st.markdown('<div class="slbl">Lease Details</div>', unsafe_allow_html=True)
    c1,c2,c3 = st.columns([2,1,1])
    with c1: lease_name = st.text_input("Lease Name", placeholder="e.g. Gachibowli Office – 3rd Floor")
    with c2: lease_from = st.text_input("Lease Start Date", placeholder="DD-MM-YYYY")
    with c3: lease_to   = st.text_input("Lease End Date",   placeholder="DD-MM-YYYY")
    c4,c5,c6 = st.columns(3)
    with c4: annual_rate     = st.number_input("Annual Discount Rate (%)", min_value=0.01,max_value=50.0,value=11.0,step=0.01,format="%.2f")
    with c5: months_override = st.number_input("No. of Months  (0 = auto from dates)", min_value=0,value=0,step=1)
    with c6: st.markdown("")

    st.markdown("---")

    # ── Section 2: Payment Details ─────────────────────────────────────────────────
    st.markdown('<div class="slbl">Payment Details</div>', unsafe_allow_html=True)

    pa,pb,pc,pd_ = st.columns([1,1,1,1])
    with pa:
        lease_payment = st.number_input(
            "Lease Payment (₹)",
            min_value=0.0, value=0.0, step=1000.0, format="%.2f",
            help="The actual cash amount paid each time (e.g. 50000 per 6-month block, not the monthly equivalent).")
    with pb:
        pay_freq = st.number_input(
            "Payment Frequency (months)",
            min_value=1, max_value=120, value=1, step=1,
            help="How often the payment is made.\n1 = Monthly | 3 = Quarterly | 6 = Half-Yearly | 12 = Yearly\nYou can enter any number of months.")
    with pc:
        freq_hint = FREQ_LABEL.get(pay_freq, f"Every {pay_freq} months")
        st.markdown(f"<br><span style='font-size:13px;color:#1D9E75;font-weight:600'>{freq_hint}</span>",
                    unsafe_allow_html=True)
    with pd_:
        advance_sel = st.selectbox("Paid in Advance?",
                                   ["No – Arrears (end of period)", "Yes – Advance (start of period)"])
        advance = advance_sel.startswith("Yes")

    # Callout
    if lease_payment > 0:
        if advance:
            st.markdown(
                f'<div class="camber">⚠️ <b>Advance – {freq_hint}:</b> '
                f'₹{lease_payment:,.2f} paid at the <b>start</b> of every {pay_freq}-month block. '
                f'First payment discount factor = <b>1.000</b> (t=0). '
                f'Excel will have 12 columns including "Advance Cash Paid".</div>',
                unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div class="callout">ℹ️ <b>Arrears – {freq_hint}:</b> '
                f'₹{lease_payment:,.2f} paid at the <b>end</b> of every {pay_freq}-month block. '
                f'Excel will have 11 columns.</div>',
                unsafe_allow_html=True)

    st.markdown("---")

    # ── Section 3: Escalation ──────────────────────────────────────────────────────
    st.markdown('<div class="slbl">Escalation Settings</div>', unsafe_allow_html=True)
    ec1,ec2 = st.columns(2)
    with ec1:
        esc_rate = st.number_input("Escalation Rate (%)", min_value=0.0,max_value=100.0,
                                   value=0.0,step=0.01,format="%.2f",
                                   help="% increase at each escalation step. 0 = no escalation.")
    with ec2:
        esc_every_n = st.number_input("Escalation Every N Months", min_value=1,max_value=360,
                                      value=12,step=1,
                                      help="Payment steps up every N months. 12=annual | 6=semi-annual | 3=quarterly | 1=monthly")
    if esc_rate > 0 and lease_payment > 0:
        after = round(lease_payment*(1+esc_rate/100),2)
        st.markdown(f'<div class="callout">💡 Payment increases by <b>{esc_rate:.2f}%</b> every '
                    f'<b>{esc_every_n} month(s)</b>: ₹{lease_payment:,.2f} → ₹{after:,.2f}.</div>',
                    unsafe_allow_html=True)

    st.markdown("")
    bcol,_ = st.columns([1,4])
    with bcol:
        do_calc = st.button("⚡  Calculate Schedule", use_container_width=True)

    # ── Calculation ────────────────────────────────────────────────────────────────
    if do_calc:
        errs=[]
        if not lease_name.strip(): errs.append("Lease name is required.")
        if lease_payment <= 0:     errs.append("Lease payment must be > 0.")
        from_date=to_date=None
        if lease_from.strip():
            try:    from_date=parse_date(lease_from)
            except ValueError as e: errs.append(str(e))
        else: from_date=date.today()
        if lease_to.strip():
            try:    to_date=parse_date(lease_to)
            except ValueError as e: errs.append(str(e))
        if months_override==0 and to_date is None:
            errs.append("Enter lease end date or number of months.")
        for e in errs: st.error(e)
        if errs: st.stop()

        eff_to = to_date if to_date else add_months(from_date, int(months_override))
        try:
            sched,pv,mr,nm,rou_d = compute_schedule(
                from_date,eff_to,lease_payment,annual_rate,
                months_override,esc_rate,esc_every_n,pay_freq,advance)
        except Exception as e:
            st.error(f"Calculation error: {e}"); st.stop()

        freq_label = FREQ_LABEL.get(pay_freq, f"Every {pay_freq} months")
        st.session_state.update({
            "sched":sched,"pv":pv,"mr":mr,"nm":nm,"rou_d":rou_d,
            "info":{
                "name":lease_name.strip() or "Lease",
                "from_date":from_date,"to_date":eff_to,
                "num_months":nm,"annual_rate":annual_rate,"monthly_rate":mr,
                "pv":pv,"lease_payment":lease_payment,
                "pay_freq":pay_freq,"advance":advance,
                "freq_label":freq_label,
                "adv_label":f"{'Advance' if advance else 'Arrears'} – {freq_label}",
                "esc_rate":esc_rate,"esc_every_n":esc_every_n,"rou_depr":rou_d,
            }
        })

    # ── Results ────────────────────────────────────────────────────────────────────
    if "sched" in st.session_state:
        sched=st.session_state["sched"]; pv=st.session_state["pv"]
        mr=st.session_state["mr"];       nm=st.session_state["nm"]
        rou_d=st.session_state["rou_d"]; info=st.session_state["info"]
        adv=info["advance"];              freq=info["pay_freq"]

        st.markdown("---")
        m1,m2,m3,m4,m5,m6=st.columns(6)
        m1.metric("Lease Term",              f"{nm} months")
        m2.metric("Annual Discount Rate",    f"{info['annual_rate']:.2f}%")
        m3.metric("Monthly Discount Rate",   f"{mr*100:.4f}%")
        m4.metric("Initial Lease Liability", inr(pv))
        m5.metric("Monthly Depreciation",    inr(rou_d))
        m6.metric("Payment Mode",            info["adv_label"])
        st.markdown("")

        tab1,tab2,tab3,tab4=st.tabs([
            "  📊 Amortisation Schedule  ",
            "  💸 Cash Payment Schedule  ",
            "  📖 Journal Entries (Yr 1)  ",
            "  📋 Payment Step Summary  ",
        ])

        # ── Tab 1 ─────────────────────────────────────────────────────────────────
        with tab1:
            rows_d=[]
            for r in sched:
                row_d={"S.No":r["sno"],"Month":r["date"].strftime("%d-%b-%Y")}
                if adv:
                    row_d["Advance Cash Paid (₹)"]=f"{r['cash']:,.2f}" if r["cash"] else "—"
                else:
                    row_d["Lease Payment (₹)"]=f"{r['cash']:,.2f}" if r["cash"] else "—"
                row_d["Discount Factor"]=f"{r['df']:.8f}" if r["df"] is not None else "—"
                row_d["Present Value (₹)"]=f"{r['pv_col']:,.2f}" if r["pv_col"] else "—"
                row_d.update({
                    "Opening Liability (₹)":f"{r['open']:,.2f}",
                    "Interest Expense (₹)": f"{r['int']:,.2f}",
                    "Closing Liability (₹)":f"{r['close']:,.2f}",
                    "Depreciation (₹)":     f"{r['depr']:,.2f}",
                    "ROU Asset (₹)":        f"{r['rou']:,.2f}",
                })
                rows_d.append(row_d)
            st.dataframe(pd.DataFrame(rows_d),use_container_width=True,hide_index=True,height=430)
            pay_rows=[r for r in sched if r["cash"]>0]
            st.caption(
                f"Rows: {nm}  ·  Payment events: {len(pay_rows)}  ·  "
                f"Total cash paid: {inr(sum(r['cash'] for r in sched))}  ·  "
                f"Total interest: {inr(sum(r['int'] for r in sched))}  ·  "
                f"Total depreciation: {inr(sum(r['depr'] for r in sched))}")

        # ── Tab 2 ─────────────────────────────────────────────────────────────────
        with tab2:
            cash_rows=[r for r in sched if r["cash"]>0]
            df_cash=[]
            for r in cash_rows:
                blk_start=r["sno"]
                blk_end  =min(r["sno"]+freq-1,nm) if adv else r["sno"]
                df_cash.append({
                    "S.No":r["sno"],
                    "Payment Date":r["date"].strftime("%d-%b-%Y"),
                    "Period":f"Month {blk_start}{'–'+str(blk_end) if blk_end!=blk_start else ''}",
                    "Cash Paid (₹)":f"{r['cash']:,.2f}",
                    "Discount Factor":f"{r['df']:.8f}" if r["df"] is not None else "—",
                    "Present Value (₹)":f"{r['pv_col']:,.2f}",
                })
            st.markdown("**Actual cash outflows with discount factors**")
            st.dataframe(pd.DataFrame(df_cash),use_container_width=True,hide_index=True,
                         height=min(450,len(df_cash)*38+50))
            st.caption(f"Total payment events: {len(cash_rows)}  ·  "
                       f"Total: {inr(sum(r['cash'] for r in cash_rows))}")

        # ── Tab 3 ─────────────────────────────────────────────────────────────────
        with tab3:
            yr1=sched[:min(12,nm)]
            yr1_d=sum(r["depr"] for r in yr1); yr1_i=sum(r["int"] for r in yr1)
            yr1_p=sum(r["cash"] for r in yr1)
            st.markdown("**Initial Recognition (Day 1)**")
            st.dataframe(pd.DataFrame([
                {"Particulars":"ROU Asset A/c","Dr (₹)":f"{pv:,.2f}","Cr (₹)":"—"},
                {"Particulars":"   To Lease Liability A/c","Dr (₹)":"—","Cr (₹)":f"{pv:,.2f}"},
            ]),hide_index=True,use_container_width=True,height=100)
            st.markdown("**Monthly entries – Year 1**")
            jrows=[]
            for r in yr1:
                d=r["date"].strftime("%d-%b-%Y")
                if r["cash"]>0 and adv:
                    jrows+=[
                        {"Date":d,"Particulars":"Lease Liability A/c  Dr","Dr (₹)":f"{r['cash']:,.2f}","Cr (₹)":"—","Narration":"Advance cash paid"},
                        {"Date":"","Particulars":"   To Bank / Cash A/c","Dr (₹)":"—","Cr (₹)":f"{r['cash']:,.2f}","Narration":""},
                    ]
                jrows+=[
                    {"Date":d,"Particulars":"Interest on Lease Liability A/c  Dr","Dr (₹)":f"{r['int']:,.2f}","Cr (₹)":"—","Narration":"Interest accrual"},
                    {"Date":"","Particulars":"   To Lease Liability A/c","Dr (₹)":"—","Cr (₹)":f"{r['int']:,.2f}","Narration":""},
                    {"Date":d,"Particulars":"Depreciation A/c  Dr","Dr (₹)":f"{r['depr']:,.2f}","Cr (₹)":"—","Narration":"ROU depreciation"},
                    {"Date":"","Particulars":"   To ROU Asset A/c","Dr (₹)":"—","Cr (₹)":f"{r['depr']:,.2f}","Narration":""},
                    {"Date":"·","Particulars":"·","Dr (₹)":"","Cr (₹)":"","Narration":""},
                ]
                if not adv and r["cash"]>0:
                    jrows.insert(-1,{"Date":d,"Particulars":"Lease Liability A/c  Dr","Dr (₹)":f"{r['cash']:,.2f}","Cr (₹)":"—","Narration":"Lease payment"})
                    jrows.insert(-1,{"Date":"","Particulars":"   To Bank / Cash A/c","Dr (₹)":"—","Cr (₹)":f"{r['cash']:,.2f}","Narration":""})
            st.dataframe(pd.DataFrame(jrows),hide_index=True,use_container_width=True,height=430)
            last_yr1=sched[min(11,nm-1)]
            st.markdown("**Year 1 Summary**")
            st.dataframe(pd.DataFrame([
                {"Particulars":"Total Depreciation (Yr 1)",    "Amount (₹)":f"{yr1_d:,.2f}"},
                {"Particulars":"Total Interest Expense (Yr 1)","Amount (₹)":f"{yr1_i:,.2f}"},
                {"Particulars":"Total Cash Paid (Yr 1)",       "Amount (₹)":f"{yr1_p:,.2f}"},
                {"Particulars":"Closing Lease Liability",      "Amount (₹)":f"{last_yr1['close']:,.2f}"},
                {"Particulars":"Closing ROU Asset",            "Amount (₹)":f"{last_yr1['rou']:,.2f}"},
            ]),hide_index=True,use_container_width=True,height=215)

        # ── Tab 4 ─────────────────────────────────────────────────────────────────
        with tab4:
            pay_rows=[r for r in sched if r["cash"]>0]
            steps=[]
            for r in pay_rows:
                steps.append({
                    "S.No":r["sno"],
                    "Payment Date":r["date"].strftime("%d-%b-%Y"),
                    "Cash Amount (₹)":f"{r['cash']:,.2f}",
                    "Discount Factor":f"{r['df']:.8f}" if r["df"] is not None else "—",
                    "PV (₹)":f"{r['pv_col']:,.2f}",
                })
            st.markdown("**All payment events across the lease term**")
            st.dataframe(pd.DataFrame(steps),hide_index=True,use_container_width=True)
            st.caption(f"Frequency: {info['adv_label']}  ·  "
                       f"Total events: {len(steps)}  ·  "
                       f"Total cash: {inr(sum(r['cash'] for r in pay_rows))}")

        # ── Download ───────────────────────────────────────────────────────────────
        st.markdown("---")
        dcol,icol=st.columns([1,3])
        with dcol:
            xl=build_excel(sched,info)
            fname=f"IndAS116_{info['name'].replace(' ','_')}.xlsx"
            st.download_button(label="⬇️  Download Excel",data=xl,file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with icol:
            cols_n="11 columns" if not adv else "12 columns (incl. Advance Cash Paid)"
            st.caption(f"Excel: {cols_n}  ·  all cells use live Excel formulas  ·  "
                       "blue=inputs · formula cells=auto-calculated · green=PV  ·  "
                       "non-payment rows show blank DF/PV as per Ind AS 116")


